from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from typing import Any

import pandas as pd
import streamlit as st

from modules.daily_report_auth import check_daily_report_password
from modules.daily_report_storage import (
    clear_current_daily_closing,
    delete_daily_closing,
    list_daily_closings,
    load_current_daily_closing,
    save_daily_closing,
)


COLUMN_ALIASES = {
    "ticket": [
        "ticket", "ticket#", "os", "ordem de servico", "ordem de serviço",
        "numero os", "número os", "id chamado", "chamado",
    ],
    "inep": ["inep", "codigo inep", "código inep", "cod inep"],
    "school": ["escola", "nome da escola", "unidade", "unidade escolar"],
    "uf": ["uf", "estado"],
    "provider": [
        "provedor", "fornecedor", "fornecedor de rede externa", "operadora", "isp",
    ],
    "days": [
        "dias", "dias em aberto", "dias aberto", "tempo em aberto",
        "dias abertos", "dias abertos corridos", "dias abertos (corridos)",
    ],
    "opened_at": [
        "aberto em", "aberto no dia", "aberto dia", "data de abertura", "data abertura", "dt abertura",
        "data/hora abertura", "data hora abertura", "abertura", "abertura do chamado",
        "criado em", "criado", "data de criacao", "data de criação",
        "dt criacao", "dt criação", "created at", "created", "opened at", "opened",
    ],
    "analyst": ["analista", "responsavel", "responsável", "atribuido a", "atribuído a"],
}

REQUIRED_FIELDS = {
    "ticket": "Ticket/OS",
    "uf": "UF/Estado",
    "provider": "Provedor/Operadora",
    "days": "Dias em aberto",
}

DISPLAY_COLUMNS = {
    "ticket": "Ticket/OS",
    "inep": "INEP",
    "school": "Escola",
    "uf": "UF",
    "provider": "Provedor",
    "days": "Dias em aberto",
    "opened_at": "Data de abertura",
    "analyst": "Analista",
}

INDICATOR_LABELS = {
    "total_open": "Total de chamados em aberto",
    "over_30": "Chamados acima de 30 dias",
    "over_60": "Chamados acima de 60 dias",
    "over_100": "Chamados acima de 100 dias",
    "affected_providers": "Total de provedores afetados",
    "opened_today": "Chamados abertos hoje",
}


def _normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_multiline_items(text: str) -> list[str]:
    raw_items = re.split(r"[\n,;]+", str(text or ""))
    return [item.strip() for item in raw_items if item.strip()]


def find_column_map(columns: list[str]) -> tuple[dict[str, str], list[str]]:
    normalized_columns = {_normalize_text(col): col for col in columns}
    found = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            key = _normalize_text(alias)
            if key in normalized_columns:
                found[field] = normalized_columns[key]
                break
        if field not in found:
            for normalized_col, original_col in normalized_columns.items():
                if any(_normalize_text(alias) in normalized_col for alias in aliases):
                    found[field] = original_col
                    break
    missing = [label for field, label in REQUIRED_FIELDS.items() if field not in found]
    return found, missing


def _to_number(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", ".", regex=False)
        .str.extract(r"(-?\d+(?:\.\d+)?)", expand=False)
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0).astype(int)


def _parse_date_series(series: pd.Series) -> pd.Series:
    raw = series.astype(str).str.strip()
    iso_mask = raw.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}")

    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    if iso_mask.any():
        parsed.loc[iso_mask] = pd.to_datetime(series[iso_mask], errors="coerce")
    if (~iso_mask).any():
        parsed.loc[~iso_mask] = pd.to_datetime(series[~iso_mask], errors="coerce", dayfirst=True)

    missing = parsed.isna()

    if missing.any():
        parsed_without_dayfirst = pd.to_datetime(series[missing], errors="coerce")
        parsed.loc[missing] = parsed_without_dayfirst
        missing = parsed.isna()

    if missing.any():
        numeric = pd.to_numeric(series[missing], errors="coerce")
        numeric_dates = pd.to_datetime(
            numeric,
            errors="coerce",
            unit="D",
            origin="1899-12-30",
        )
        parsed.loc[missing] = numeric_dates

    return parsed.dt.date


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date()


def _records(df: pd.DataFrame) -> list[dict]:
    if df is None or df.empty:
        return []
    return df.fillna("").to_dict(orient="records")


def _indicators_dataframe(indicators: dict) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Indicador": label, "Quantidade": indicators.get(key, 0)}
            for key, label in INDICATOR_LABELS.items()
        ]
    )


def read_daily_sheet(uploaded_file) -> tuple[pd.DataFrame | None, dict[str, str], list[str], str | None]:
    if uploaded_file is None:
        return None, {}, [], "Envie uma planilha .xlsx para concluir o fechamento."

    file_name = getattr(uploaded_file, "name", "")
    if not str(file_name).lower().endswith(".xlsx"):
        return None, {}, [], "Formato inválido. Envie apenas arquivo .xlsx."

    try:
        df = pd.read_excel(uploaded_file)
    except Exception as exc:
        return None, {}, [], f"Não foi possível ler a planilha: {exc}"

    if df.empty:
        return None, {}, [], "A planilha enviada está vazia."

    column_map, missing = find_column_map([str(c) for c in df.columns])
    if missing:
        return None, column_map, missing, "Colunas obrigatórias não encontradas: " + ", ".join(missing)

    return normalize_daily_dataframe(df, column_map), column_map, [], None


def normalize_daily_dataframe(df: pd.DataFrame, column_map: dict[str, str]) -> pd.DataFrame:
    result = pd.DataFrame()
    for field, label in DISPLAY_COLUMNS.items():
        if field in column_map:
            result[label] = df[column_map[field]]
        else:
            result[label] = ""

    result["Ticket/OS"] = result["Ticket/OS"].astype(str).str.strip()
    result["INEP"] = result["INEP"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    result["UF"] = result["UF"].astype(str).str.strip().str.upper()
    result["Provedor"] = result["Provedor"].astype(str).str.strip()
    result["Dias em aberto"] = _to_number(result["Dias em aberto"])

    if "opened_at" in column_map:
        result["Data de abertura"] = _parse_date_series(result["Data de abertura"])
    elif "Data de abertura" in result.columns:
        result["Data de abertura"] = pd.NaT

    extra_cols = [c for c in df.columns if c not in set(column_map.values())]
    for col in extra_cols:
        result[str(col)] = df[col]

    return result


def _opened_today(df: pd.DataFrame, report_date: date) -> pd.DataFrame:
    if "Data de abertura" not in df.columns:
        return pd.DataFrame()
    target_date = _coerce_date(report_date)
    if target_date is None:
        target_date = date.today()
    opened_dates = _parse_date_series(df["Data de abertura"])
    return df[opened_dates == target_date].copy()


def _provider_summary(df: pd.DataFrame) -> pd.DataFrame:
    valid = df.copy()
    valid["Provedor"] = valid["Provedor"].replace({"": "Não identificado", "nan": "Não identificado"})
    summary = (
        valid.groupby("Provedor", dropna=False)
        .agg(
            Quantidade=("Ticket/OS", "count"),
            **{
                "Média de dias em aberto": ("Dias em aberto", "mean"),
                "Maior tempo em aberto": ("Dias em aberto", "max"),
            },
        )
        .reset_index()
        .sort_values(["Quantidade", "Maior tempo em aberto"], ascending=False)
    )
    summary["Média de dias em aberto"] = summary["Média de dias em aberto"].round(1)
    return summary


def _uf_summary(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    summary = (
        df.groupby("UF", dropna=False)
        .agg(Quantidade=("Ticket/OS", "count"))
        .reset_index()
        .sort_values("Quantidade", ascending=False)
    )
    summary["UF"] = summary["UF"].replace({"": "Não informado", "nan": "Não informado"})
    summary["Percentual"] = summary["Quantidade"].map(lambda x: f"{(x / total * 100):.1f}%" if total else "0%")
    return summary


def _automatic_alerts(df: pd.DataFrame, provider_summary: pd.DataFrame, uf_summary: pd.DataFrame) -> list[str]:
    alerts = []
    over_100 = int((df["Dias em aberto"] > 100).sum())
    if over_100:
        alerts.append(f"{over_100} chamado(s) com mais de 100 dias em aberto.")

    if not provider_summary.empty:
        top_provider = provider_summary.iloc[0]
        alerts.append(
            f"Maior concentração por provedor: {top_provider['Provedor']} "
            f"com {int(top_provider['Quantidade'])} chamado(s)."
        )

    if not uf_summary.empty:
        top_uf = uf_summary.iloc[0]
        alerts.append(f"UF com maior volume: {top_uf['UF']} com {int(top_uf['Quantidade'])} chamado(s).")

    without_provider = int(df["Provedor"].astype(str).str.strip().isin(["", "nan", "—", "-"]).sum())
    if without_provider:
        alerts.append(f"{without_provider} registro(s) sem provedor identificado.")

    empty_tickets = int(df["Ticket/OS"].astype(str).str.strip().isin(["", "nan"]).sum())
    if empty_tickets:
        alerts.append(f"{empty_tickets} registro(s) sem Ticket/OS preenchido.")

    urgent_mask = df.astype(str).apply(
        lambda row: row.str.contains("urgent|critico|crítico", case=False, regex=True).any(),
        axis=1,
    )
    urgent_count = int(urgent_mask.sum())
    if urgent_count:
        alerts.append(f"{urgent_count} chamado(s) com indicação de urgência ou criticidade na planilha.")

    return alerts


def build_daily_closing(
    report_date: date,
    responsible: str,
    uploaded_file,
    updated_text: str,
    closed_text: str,
    observations: str,
    attention_points: str,
    next_actions: str,
) -> tuple[dict | None, str | None]:
    df, column_map, _, error = read_daily_sheet(uploaded_file)
    if error:
        return None, error

    updated_items = parse_multiline_items(updated_text)
    closed_items = parse_multiline_items(closed_text)
    opened_today = _opened_today(df, report_date)
    critical = df.sort_values("Dias em aberto", ascending=False).head(15)
    provider_summary = _provider_summary(df)
    uf_summary = _uf_summary(df)
    alerts = _automatic_alerts(df, provider_summary, uf_summary)
    manual_attention = parse_multiline_items(attention_points)

    payload = {
        "report_date": report_date.isoformat(),
        "report_date_display": report_date.strftime("%d/%m/%Y"),
        "responsible": responsible.strip() or "Não informado",
        "source_file": getattr(uploaded_file, "name", "planilha.xlsx"),
        "updated_items": updated_items,
        "closed_items": closed_items,
        "observations": str(observations or "").strip(),
        "attention_points": str(attention_points or "").strip(),
        "manual_attention_items": manual_attention,
        "next_actions": str(next_actions or "").strip(),
        "indicators": {
            "total_open": int(len(df)),
            "over_30": int((df["Dias em aberto"] > 30).sum()),
            "over_60": int((df["Dias em aberto"] > 60).sum()),
            "over_100": int((df["Dias em aberto"] > 100).sum()),
            "affected_providers": int(df["Provedor"].replace("", pd.NA).dropna().nunique()),
            "opened_today": int(len(opened_today)),
        },
        "alerts": alerts,
        "critical_rows": _records(critical),
        "provider_summary": _records(provider_summary),
        "uf_summary": _records(uf_summary),
        "opened_today_rows": _records(opened_today),
        "has_opening_date_column": "opened_at" in column_map,
        "full_rows": _records(df),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    return payload, None


def daily_report_markdown(closing: dict) -> str:
    data = closing.get("report_date_display", closing.get("report_date", ""))
    updated = closing.get("updated_items", [])
    closed = closing.get("closed_items", [])
    observations = closing.get("observations") or "Não foram registradas observações operacionais adicionais."
    attention_manual = closing.get("manual_attention_items", [])
    next_actions = closing.get("next_actions") or "Manter acompanhamento dos chamados críticos e pendências com as provedoras."

    def bullet(items: list[str], empty: str) -> str:
        if not items:
            return empty
        return "\n".join(f"- {item}" for item in items)

    indicators = closing.get("indicators", {})
    alerts = closing.get("alerts", []) + attention_manual

    return f"""# Relatório Diário de Monitoramento - Rede Externa

Responsável: {closing.get('responsible', 'Não informado')}
Data: {data}

## 1. Resumo das Atividades do Dia

No dia {data}, foram realizadas atividades de acompanhamento, atualização e tratamento dos chamados de Rede Externa vinculados ao projeto. Foram atualizados os seguintes chamados no sistema:

{bullet(updated, "Não foram informados chamados atualizados para esta data.")}

Também foram encerrados os seguintes chamados:

{bullet(closed, "Não foram informados chamados encerrados para esta data.")}

Além disso, foi mantido contato e monitoramento junto às provedoras responsáveis. Os casos pendentes permanecem aguardando atuação das operadoras, com prioridade para chamados classificados como urgentes e com maior tempo em aberto.

## 2. Indicadores Gerais

- Total de chamados em aberto: {indicators.get('total_open', 0)}
- Chamados acima de 30 dias: {indicators.get('over_30', 0)}
- Chamados acima de 60 dias: {indicators.get('over_60', 0)}
- Chamados acima de 100 dias: {indicators.get('over_100', 0)}
- Provedores afetados: {indicators.get('affected_providers', 0)}
- Chamados abertos hoje: {indicators.get('opened_today', 0)}

## 3. Pontos de Atenção

{bullet(alerts, "Não há pontos de atenção adicionais registrados.")}

## 4. Observações Operacionais

{observations}

## 5. Próximas Ações

{next_actions}
"""


def _pdf_paragraph(text: Any, style):
    from reportlab.platypus import Paragraph

    clean = str(text if text is not None else "").replace("\n", "<br/>")
    return Paragraph(clean, style)


def _pdf_table(data: list[list[Any]], widths=None, header=True):
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Table, TableStyle

    cell_style = ParagraphStyle(
        "HubTableCell",
        fontName="Helvetica",
        fontSize=7.2,
        leading=9,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "HubTableHeader",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    wrapped_data = []
    for row_idx, row in enumerate(data):
        style = header_style if header and row_idx == 0 else cell_style
        wrapped_data.append([_pdf_paragraph(str(value), style) for value in row])

    table = Table(wrapped_data, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    style = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE9")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        style.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
        if len(data) > 1:
            style.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]))
    table.setStyle(TableStyle(style))
    return table


def _pdf_section(title: str, styles):
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph, Table, TableStyle

    table = Table([[Paragraph(title, styles["SectionTitle"])]], colWidths=[520])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E5E7EB")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _pdf_bullets(items: list[str], empty_message: str, styles):
    from reportlab.platypus import ListFlowable, ListItem, Paragraph

    if not items:
        return [Paragraph(empty_message, styles["Body"])]
    return [
        ListFlowable(
            [ListItem(Paragraph(str(item), styles["Body"]), leftIndent=12) for item in items],
            bulletType="bullet",
            start="circle",
            leftIndent=18,
        )
    ]


def _truncate_table_value(value: Any, limit: int = 70) -> str:
    text = str(value if value is not None else "")
    text = re.sub(r"\s+", " ", text).strip()
    return text[: limit - 3] + "..." if len(text) > limit else text


def generate_daily_report_pdf(closing: dict) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=1.15 * cm,
        bottomMargin=1.15 * cm,
        title="Relatório Diário de Monitoramento - Rede Externa",
    )

    base = getSampleStyleSheet()
    styles = {
        "Title": ParagraphStyle(
            "HubTitle",
            parent=base["Title"],
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#111827"),
            spaceAfter=4,
        ),
        "Subtitle": ParagraphStyle(
            "HubSubtitle",
            parent=base["Normal"],
            alignment=TA_CENTER,
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#475569"),
            spaceAfter=12,
        ),
        "SectionTitle": ParagraphStyle(
            "HubSectionTitle",
            parent=base["Normal"],
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#111827"),
        ),
        "Body": ParagraphStyle(
            "HubBody",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#111827"),
            spaceAfter=7,
        ),
        "Small": ParagraphStyle(
            "HubSmall",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#475569"),
        ),
    }

    story = [
        Paragraph("Relatório Diário de Monitoramento - Rede Externa", styles["Title"]),
        Paragraph("Hub Redes - EACE", styles["Subtitle"]),
    ]

    report_date = closing.get("report_date_display", closing.get("report_date", ""))
    metadata = [
        ["Data do relatório", report_date, "Responsável", closing.get("responsible", "Não informado")],
        ["Fonte", closing.get("source_file", "planilha.xlsx"), "Gerado em", datetime.now().strftime("%d/%m/%Y")],
    ]
    story.append(_pdf_table(metadata, widths=[3.2 * cm, 6.0 * cm, 3.0 * cm, 6.0 * cm], header=False))
    story.append(Spacer(1, 10))

    story.append(_pdf_section("1. Resumo das Atividades do Dia", styles))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            f"No dia {report_date}, foram realizadas atividades de acompanhamento, atualização e tratamento "
            "dos chamados de Rede Externa vinculados ao projeto.",
            styles["Body"],
        )
    )
    story.append(Paragraph("<b>Chamados/INEPs atualizados no sistema:</b>", styles["Body"]))
    story.extend(_pdf_bullets(closing.get("updated_items", []), "Não foram informados chamados atualizados para esta data.", styles))
    story.append(Paragraph("<b>Chamados/INEPs encerrados:</b>", styles["Body"]))
    story.extend(_pdf_bullets(closing.get("closed_items", []), "Não foram informados chamados encerrados para esta data.", styles))
    story.append(
        Paragraph(
            "Além disso, foi mantido contato e monitoramento junto às provedoras responsáveis. "
            "Os casos pendentes permanecem aguardando atuação das operadoras, com prioridade para chamados "
            "classificados como urgentes e com maior tempo em aberto.",
            styles["Body"],
        )
    )

    story.append(_pdf_section("2. Indicadores Gerais", styles))
    story.append(Spacer(1, 6))
    indicator_rows = [["Indicador", "Quantidade"]] + _indicators_dataframe(closing.get("indicators", {})).values.tolist()
    story.append(_pdf_table(indicator_rows, widths=[12.5 * cm, 4.0 * cm]))
    story.append(Spacer(1, 10))

    story.append(_pdf_section("3. Chamados Mais Críticos", styles))
    story.append(Spacer(1, 6))
    critical_df = pd.DataFrame(closing.get("critical_rows", []))
    if critical_df.empty:
        story.append(Paragraph("Não há chamados criticos para exibir.", styles["Body"]))
    else:
        cols = [c for c in ["Ticket/OS", "INEP", "Escola", "UF", "Provedor", "Dias em aberto"] if c in critical_df.columns]
        critical_df = critical_df[cols].head(15).astype(str)
        rows = [cols] + critical_df.map(lambda x: _truncate_table_value(x, 58)).values.tolist()
        story.append(_pdf_table(rows, widths=[2.7 * cm, 2.2 * cm, 6.0 * cm, 1.1 * cm, 4.0 * cm, 2.2 * cm]))
    story.append(Spacer(1, 10))

    story.append(_pdf_section("4. Chamados por Provedor", styles))
    story.append(Spacer(1, 6))
    provider_df = pd.DataFrame(closing.get("provider_summary", []))
    if provider_df.empty:
        story.append(Paragraph("não h? dados por provedor.", styles["Body"]))
    else:
        cols = provider_df.columns.tolist()
        rows = [cols] + provider_df.head(20).astype(str).map(lambda x: _truncate_table_value(x, 52)).values.tolist()
        story.append(_pdf_table(rows, widths=[7.2 * cm, 2.8 * cm, 4.0 * cm, 3.4 * cm]))
    story.append(Spacer(1, 10))

    story.append(_pdf_section("5. Distribuicão por Estado", styles))
    story.append(Spacer(1, 6))
    uf_df = pd.DataFrame(closing.get("uf_summary", []))
    if uf_df.empty:
        story.append(Paragraph("não h? dados por UF.", styles["Body"]))
    else:
        rows = [uf_df.columns.tolist()] + uf_df.astype(str).values.tolist()
        story.append(_pdf_table(rows, widths=[5.0 * cm, 5.5 * cm, 5.5 * cm]))
    story.append(Spacer(1, 10))

    story.append(_pdf_section("6. Todos os Chamados de Rede Externa", styles))
    story.append(Spacer(1, 6))
    full_df = pd.DataFrame(closing.get("full_rows", []))
    if full_df.empty:
        story.append(Paragraph("não h? listagem completa dispon?vel para este fechamento.", styles["Body"]))
    else:
        cols = [c for c in ["Ticket/OS", "Provedor", "Dias em aberto", "INEP", "Escola"] if c in full_df.columns]
        full_df = full_df[cols].astype(str)
        rows = [cols] + full_df.map(lambda x: _truncate_table_value(x, 68)).values.tolist()
        story.append(_pdf_table(rows, widths=[2.8 * cm, 4.3 * cm, 2.4 * cm, 2.3 * cm, 6.6 * cm]))
    story.append(Spacer(1, 10))

    story.append(_pdf_section("7. Pontos de Atenção", styles))
    story.append(Spacer(1, 6))
    story.extend(_pdf_bullets(
        closing.get("alerts", []) + closing.get("manual_attention_items", []),
        "Não há pontos de Atenção adicionais registrados.",
        styles,
    ))

    story.append(_pdf_section("8. Observações Operacionais", styles))
    story.append(Spacer(1, 6))
    story.append(Paragraph(closing.get("observations") or "não foram registradas Observações operacionais adicionais.", styles["Body"]))

    story.append(_pdf_section("9. Proximas atividades", styles))
    story.append(Spacer(1, 6))
    story.append(Paragraph(closing.get("next_actions") or "Manter acompanhamento dos chamados críticos e pendências com as provedoras.", styles["Body"]))

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(1.25 * cm, 0.7 * cm, "Hub Redes - EACE")
        canvas.drawRightString(A4[0] - 1.25 * cm, 0.7 * cm, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


def generate_daily_report_excel(closing: dict) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils.cell import range_boundaries

    colors = {
        "navy": "1F3864",
        "blue": "2E75B6",
        "red": "C00000",
        "light_red": "FFCCCC",
        "light_orange": "FDEBD0",
        "light_yellow": "FFFACD",
        "light_blue": "D6E4F0",
        "light_green": "E2EFDA",
        "next_blue": "BDD7EE",
        "white": "FFFFFF",
        "black": "000000",
    }
    thin = Side(style="thin", color="A6A6A6")
    medium = Side(style="medium", color="1F3864")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    def fill(name: str) -> PatternFill:
        return PatternFill("solid", fgColor=colors[name])

    def font(size=10, bold=False, color="black") -> Font:
        return Font(name="Arial", size=size, bold=bold, color=colors[color])

    def align(horizontal="center", vertical="center", wrap=False) -> Alignment:
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

    def style_cell(cell, *, fill_name=None, font_obj=None, alignment=None, border=None, number_format=None):
        if fill_name:
            cell.fill = fill(fill_name)
        if font_obj:
            cell.font = font_obj
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format

    def style_range(ws, cell_range: str, *, fill_name=None, font_obj=None, alignment=None, border=None):
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                style_cell(
                    cell,
                    fill_name=fill_name,
                    font_obj=font_obj,
                    alignment=alignment,
                    border=border,
                )

    def merge_set(ws, cell_range: str, value, *, fill_name=None, font_obj=None, alignment=None, border=None):
        style_range(ws, cell_range, fill_name=fill_name, font_obj=font_obj, alignment=alignment, border=border)
        ws.merge_cells(cell_range)
        top_left = ws[cell_range.split(":")[0]]
        top_left.value = value

    def safe_text(value: Any) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except TypeError:
            pass
        return str(value)

    def days_value(row: dict) -> int:
        try:
            return int(float(row.get("Dias em aberto", 0) or 0))
        except (TypeError, ValueError):
            return 0

    def severity_fill(days: int, index: int) -> str:
        if days > 100:
            return "light_red"
        if days > 60:
            return "light_orange"
        if days > 30:
            return "light_yellow"
        return "light_blue" if index % 2 == 0 else "white"

    def percent_value(value: Any) -> float:
        text = safe_text(value).replace("%", "").replace(",", ".").strip()
        try:
            number = float(text)
        except ValueError:
            return 0.0
        return number / 100 if number > 1 else number

    def prepare_sheet(ws, widths: dict[str, float]):
        ws.sheet_view.showGridLines = False
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.page_margins.left = 0.75
        ws.page_margins.right = 0.75
        ws.page_margins.top = 1.0
        ws.page_margins.bottom = 1.0

    report_date = closing.get("report_date_display", closing.get("report_date", ""))
    generated_date = datetime.now().strftime("%d/%m/%Y")
    indicators = closing.get("indicators", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Executivo"
    prepare_sheet(
        ws,
        {"A": 3, "B": 22, "C": 18, "D": 13, "E": 13, "F": 13, "G": 13, "H": 13, "I": 13, "J": 13},
    )
    for row, height in {
        1: 7.5, 2: 37.5, 3: 24, 4: 7.5, 5: 21.75, 6: 9.75, 7: 25.5,
        8: 60, 9: 27.75, 10: 12, 11: 25.5, 12: 19.5,
    }.items():
        ws.row_dimensions[row].height = height

    merge_set(ws, "B2:J2", "HUB REDES \u2013 EACE", fill_name="navy", font_obj=font(18, True, "white"), alignment=align())
    merge_set(
        ws,
        "B3:J3",
        "RELAT\u00d3RIO DI\u00c1RIO DE MONITORAMENTO \u2013 REDE EXTERNA",
        fill_name="blue",
        font_obj=font(12, True, "white"),
        alignment=align(),
    )
    metadata = {
        "B5": "Data do Relat\u00f3rio:", "C5": report_date,
        "E5": "Respons\u00e1vel:", "F5": closing.get("responsible", "N\u00e3o informado"),
        "H5": "Gerado em:", "I5": generated_date,
    }
    for coord, value in metadata.items():
        ws[coord] = value
        is_label = coord[0] in {"B", "E", "H"}
        style_cell(
            ws[coord],
            font_obj=font(10, is_label, "navy" if is_label else "black"),
            alignment=align("right" if is_label else "left"),
        )

    merge_set(ws, "B7:J7", "INDICADORES GERAIS", fill_name="navy", font_obj=font(11, True, "white"), alignment=align(), border=medium_border)
    cards = [
        ("B", "C", indicators.get("total_open", 0), "Total de Chamados em Aberto"),
        ("D", "E", indicators.get("over_30", 0), "Chamados Acima de 30 Dias"),
        ("F", "G", indicators.get("over_60", 0), "Chamados Acima de 60 Dias"),
        ("H", "I", indicators.get("over_100", 0), "Chamados Acima de 100 Dias"),
    ]
    for start_col, end_col, value, label in cards:
        merge_set(ws, f"{start_col}8:{end_col}8", value, fill_name="blue", font_obj=font(28, True, "white"), alignment=align(), border=medium_border)
        merge_set(ws, f"{start_col}9:{end_col}9", label, fill_name="navy", font_obj=font(9, True, "white"), alignment=align(wrap=True), border=medium_border)

    merge_set(ws, "B11:J11", "CHAMADOS MAIS CR\u00cdTICOS (TOP 15)", fill_name="navy", font_obj=font(11, True, "white"), alignment=align(), border=medium_border)
    for coord, value in {"B12": "Ticket/OS", "C12": "INEP", "D12:F12": "Escola", "G12": "UF", "H12:I12": "Provedor", "J12": "Dias em Aberto"}.items():
        if ":" in coord:
            merge_set(ws, coord, value, fill_name="blue", font_obj=font(10, True, "white"), alignment=align(), border=thin_border)
        else:
            ws[coord] = value
            style_cell(ws[coord], fill_name="blue", font_obj=font(10, True, "white"), alignment=align(), border=thin_border)

    critical_df = pd.DataFrame(closing.get("critical_rows", []))
    if critical_df.empty:
        full_candidates = pd.DataFrame(closing.get("full_rows", []))
        if not full_candidates.empty and "Dias em aberto" in full_candidates.columns:
            critical_df = full_candidates.sort_values("Dias em aberto", ascending=False).head(15)
        else:
            critical_df = full_candidates.head(15)
    for index in range(15):
        row_num = 13 + index
        ws.row_dimensions[row_num].height = 19.5
        record = critical_df.iloc[index].to_dict() if index < len(critical_df) else {}
        row_fill = severity_fill(days_value(record), index)
        values = {
            "B": safe_text(record.get("Ticket/OS")),
            "C": safe_text(record.get("INEP")),
            "D:F": safe_text(record.get("Escola")),
            "G": safe_text(record.get("UF")),
            "H:I": safe_text(record.get("Provedor")),
            "J": days_value(record) if record else "",
        }
        for col_ref, value in values.items():
            if ":" in col_ref:
                merge_set(ws, f"{col_ref.split(':')[0]}{row_num}:{col_ref.split(':')[1]}{row_num}", value, fill_name=row_fill, font_obj=font(9), alignment=align("left", wrap=True), border=thin_border)
            else:
                ws[f"{col_ref}{row_num}"] = value
                horizontal = "center" if col_ref in {"B", "C", "G", "J"} else "left"
                style_cell(ws[f"{col_ref}{row_num}"], fill_name=row_fill, font_obj=font(9), alignment=align(horizontal, wrap=True), border=thin_border)

    merge_set(ws, "B29:J29", "CHAMADOS POR PROVEDOR", fill_name="navy", font_obj=font(11, True, "white"), alignment=align(), border=medium_border)
    provider_headers = {"B30:D30": "Provedor", "E30:F30": "Qtd. Chamados", "G30:H30": "M\u00e9dia Dias em Aberto", "I30:J30": "Maior Tempo (Dias)"}
    for cell_range, value in provider_headers.items():
        merge_set(ws, cell_range, value, fill_name="blue", font_obj=font(10, True, "white"), alignment=align(), border=thin_border)
    provider_df = pd.DataFrame(closing.get("provider_summary", [])).head(19)
    for index in range(19):
        row_num = 31 + index
        record = provider_df.iloc[index].to_dict() if index < len(provider_df) else {}
        row_fill = "light_blue" if index % 2 == 0 else "white"
        merge_set(ws, f"B{row_num}:D{row_num}", safe_text(record.get("Provedor")), fill_name=row_fill, font_obj=font(9), alignment=align("left"), border=thin_border)
        merge_set(ws, f"E{row_num}:F{row_num}", record.get("Quantidade", "") if record else "", fill_name=row_fill, font_obj=font(9), alignment=align(), border=thin_border)
        merge_set(ws, f"G{row_num}:H{row_num}", record.get("M\u00e9dia de dias em aberto", "") if record else "", fill_name=row_fill, font_obj=font(9), alignment=align(), border=thin_border)
        merge_set(ws, f"I{row_num}:J{row_num}", record.get("Maior tempo em aberto", "") if record else "", fill_name=row_fill, font_obj=font(9), alignment=align(), border=thin_border)

    uf_start = 51
    merge_set(ws, f"B{uf_start}:J{uf_start}", "DISTRIBUI\u00c7\u00c3O POR ESTADO (UF)", fill_name="navy", font_obj=font(11, True, "white"), alignment=align(), border=medium_border)
    for cell_range, value in {f"B{uf_start + 1}:C{uf_start + 1}": "UF", f"D{uf_start + 1}:F{uf_start + 1}": "Quantidade", f"G{uf_start + 1}:J{uf_start + 1}": "Percentual"}.items():
        merge_set(ws, cell_range, value, fill_name="blue", font_obj=font(10, True, "white"), alignment=align(), border=thin_border)
    uf_df = pd.DataFrame(closing.get("uf_summary", []))
    uf_rows = max(7, len(uf_df))
    for index in range(uf_rows):
        row_num = uf_start + 2 + index
        record = uf_df.iloc[index].to_dict() if index < len(uf_df) else {}
        row_fill = "light_blue" if index % 2 == 0 else "white"
        merge_set(ws, f"B{row_num}:C{row_num}", safe_text(record.get("UF")), fill_name=row_fill, font_obj=font(9), alignment=align(), border=thin_border)
        merge_set(ws, f"D{row_num}:F{row_num}", record.get("Quantidade", "") if record else "", fill_name=row_fill, font_obj=font(9), alignment=align(), border=thin_border)
        merge_set(ws, f"G{row_num}:J{row_num}", percent_value(record.get("Percentual", 0)) if record else "", fill_name=row_fill, font_obj=font(9), alignment=align(), border=thin_border)
        if record:
            ws[f"G{row_num}"].number_format = "0.0%"

    attention_start = max(61, uf_start + 3 + uf_rows)
    merge_set(ws, f"B{attention_start}:J{attention_start}", "PONTOS DE ATEN\u00c7\u00c3O", fill_name="red", font_obj=font(11, True, "white"), alignment=align(), border=medium_border)
    alerts = closing.get("alerts", []) + closing.get("manual_attention_items", [])
    if not alerts:
        alerts = ["N\u00e3o h\u00e1 pontos de aten\u00e7\u00e3o adicionais registrados."]
    for index, item in enumerate(alerts):
        row_num = attention_start + 1 + index
        merge_set(ws, f"B{row_num}:J{row_num}", f"\u2022 {item}", fill_name="light_red", font_obj=font(10), alignment=align("left", wrap=True), border=thin_border)
        ws.row_dimensions[row_num].height = 24

    next_start = attention_start + len(alerts) + 2
    merge_set(ws, f"B{next_start}:J{next_start}", "PR\u00d3XIMAS ATIVIDADES", fill_name="blue", font_obj=font(11, True, "white"), alignment=align(), border=medium_border)
    merge_set(
        ws,
        f"B{next_start + 1}:J{next_start + 1}",
        closing.get("next_actions") or "Manter acompanhamento dos chamados cr\u00edticos e pend\u00eancias com as provedoras.",
        fill_name="next_blue",
        font_obj=font(10),
        alignment=align("left", wrap=True),
        border=thin_border,
    )
    ws.row_dimensions[next_start + 1].height = 28

    ws_all = wb.create_sheet("Todos os Chamados")
    prepare_sheet(ws_all, {"A": 3, "B": 18, "C": 22, "D": 15, "E": 9, "F": 50})
    ws_all.row_dimensions[1].height = 7.5
    ws_all.row_dimensions[2].height = 31.5
    ws_all.row_dimensions[3].height = 9.75
    merge_set(
        ws_all,
        "B2:F2",
        f"HUB REDES \u2013 EACE  |  TODOS OS CHAMADOS DE REDE EXTERNA  |  {report_date}",
        fill_name="navy",
        font_obj=font(13, True, "white"),
        alignment=align(),
    )
    headers = ["Ticket/OS", "Provedor", "Dias em Aberto", "INEP", "Escola"]
    for col_idx, header in enumerate(headers, start=2):
        cell = ws_all.cell(4, col_idx, header)
        style_cell(cell, fill_name="blue", font_obj=font(10, True, "white"), alignment=align(), border=thin_border)
    full_df = pd.DataFrame(closing.get("full_rows", []))
    if not full_df.empty and "Dias em aberto" in full_df.columns:
        full_df = full_df.sort_values("Dias em aberto", ascending=False)
    for index, (_, record) in enumerate(full_df.iterrows(), start=0):
        row_num = 5 + index
        ws_all.row_dimensions[row_num].height = 16.5
        data = record.to_dict()
        row_fill = severity_fill(days_value(data), index)
        values = [safe_text(data.get("Ticket/OS")), safe_text(data.get("Provedor")), days_value(data), safe_text(data.get("INEP")), safe_text(data.get("Escola"))]
        for col_idx, value in enumerate(values, start=2):
            cell = ws_all.cell(row_num, col_idx, value)
            horizontal = "center" if col_idx == 4 else "left"
            style_cell(cell, fill_name=row_fill, font_obj=font(9), alignment=align(horizontal), border=thin_border)
    total_row = 5 + len(full_df)
    merge_set(ws_all, f"B{total_row}:D{total_row}", "TOTAL DE CHAMADOS", fill_name="navy", font_obj=font(10, True, "white"), alignment=align(), border=medium_border)
    ws_all[f"E{total_row}"] = len(full_df)
    style_cell(ws_all[f"E{total_row}"], fill_name="navy", font_obj=font(10, True, "white"), alignment=align(), border=medium_border)

    ws_mov = wb.create_sheet("Movimenta\u00e7\u00f5es do Dia")
    prepare_sheet(ws_mov, {"A": 3, "B": 72.57, "C": 34.43})
    ws_mov.row_dimensions[1].height = 7.5
    ws_mov.row_dimensions[2].height = 31.5
    ws_mov.row_dimensions[3].height = 9.75
    merge_set(
        ws_mov,
        "B2:C2",
        f"HUB REDES \u2013 EACE  |  MOVIMENTA\u00c7\u00d5ES DO DIA \u2013 {report_date}",
        fill_name="navy",
        font_obj=font(13, True, "white"),
        alignment=align(),
    )
    for coord, label in {"B4": "ATUALIZADOS NO SISTEMA", "C4": "ENCERRADOS"}.items():
        ws_mov[coord] = label
        style_cell(ws_mov[coord], fill_name="blue", font_obj=font(10, True, "white"), alignment=align(), border=medium_border)
    movements = [("B", closing.get("updated_items", []), "light_blue"), ("C", closing.get("closed_items", []), "light_green")]
    for col, items, color_name in movements:
        for index, item in enumerate(items):
            row_num = 5 + index
            ws_mov.row_dimensions[row_num].height = 18
            cell = ws_mov[f"{col}{row_num}"]
            cell.value = item
            style_cell(cell, fill_name=color_name if index % 2 == 0 else "white", font_obj=font(10), alignment=align(), border=thin_border)
        total_row = 5 + len(items)
        ws_mov[f"{col}{total_row}"] = f"Total: {len(items)}"
        style_cell(ws_mov[f"{col}{total_row}"], fill_name="navy", font_obj=font(10, True, "white"), alignment=align(), border=medium_border)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _render_item_list(items: list[str], empty_message: str) -> None:
    if items:
        for item in items:
            st.markdown(f"- {item}")
    else:
        st.info(empty_message)


def render_daily_report(closing: dict | None) -> None:
    st.markdown("### Relatório Diário de Monitoramento — Rede Externa")

    history = list_daily_closings()
    if not closing:
        closing = load_current_daily_closing()

    report_options = []
    if closing:
        report_options.append(
            {
                "label": "Relatório atual",
                "path": closing.get("history_path"),
                "data": closing,
                "is_current": True,
            }
        )
    report_options.extend({**item, "is_current": False} for item in history)

    if report_options:
        if "daily_report_history_select" in st.session_state and not isinstance(
            st.session_state.daily_report_history_select, dict
        ):
            del st.session_state.daily_report_history_select

        col_history, col_delete = st.columns([12, 1])
        with col_history:
            selected_report = st.selectbox(
                "Consultar histórico",
                report_options,
                format_func=lambda item: item["label"],
                key="daily_report_history_select",
            )
        with col_delete:
            st.markdown("<div style='height: 1.7rem'></div>", unsafe_allow_html=True)
            if st.button("🗑️", key="delete_daily_report", help="Excluir relatório selecionado", use_container_width=True):
                st.session_state.daily_report_delete_pending = selected_report

        pending_delete = st.session_state.get("daily_report_delete_pending")
        if pending_delete:
            with st.expander("Confirmar exclusão de relatório", expanded=True):
                st.warning(
                    f"Você está prestes a excluir: {pending_delete.get('label', 'relatório selecionado')}. "
                    "Essa ação remove o relatório para todos os usuários."
                )
                delete_password = st.text_input(
                    "Senha para excluir",
                    type="password",
                    key="daily_report_delete_password",
                )
                confirm_col, cancel_col = st.columns(2)
                if confirm_col.button("Confirmar exclusão", type="primary", use_container_width=True):
                    if not check_daily_report_password(delete_password):
                        st.error("Senha inválida.")
                    else:
                        ok = delete_daily_closing(pending_delete.get("path"))
                        if pending_delete.get("is_current"):
                            clear_current_daily_closing()
                        st.session_state.daily_report_current = None
                        st.session_state.daily_report_delete_pending = None
                        if "daily_report_delete_password" in st.session_state:
                            del st.session_state.daily_report_delete_password
                        if ok:
                            st.success("Relatório excluído. Ele não ficará mais acessível para outros usuários.")
                        else:
                            st.error("Não foi possível excluir o relatório selecionado.")
                        st.rerun()
                if cancel_col.button("Cancelar", use_container_width=True):
                    st.session_state.daily_report_delete_pending = None
                    if "daily_report_delete_password" in st.session_state:
                        del st.session_state.daily_report_delete_password
                    st.rerun()
        closing = selected_report["data"]

    if not closing:
        st.info("Nenhum fechamento diário concluído ainda. Use a Área Restrita na lateral para gerar o relatório.")
        return

    indicators = closing.get("indicators", {})
    st.caption(
        f"Data: {closing.get('report_date_display', closing.get('report_date', ''))} | "
        f"Responsável: {closing.get('responsible', 'Não informado')} | "
        f"Fonte: {closing.get('source_file', 'planilha.xlsx')}"
    )

    st.markdown("""
    <div class="saas-grid">
      <div class="saas-card"><div class="saas-card-label">Chamados abertos</div>
        <div class="saas-card-value c-blue">{total}</div><div class="saas-card-sub">planilha diária</div></div>
      <div class="saas-card"><div class="saas-card-label">Acima de 30 dias</div>
        <div class="saas-card-value c-yellow">{over30}</div><div class="saas-card-sub">atenção</div></div>
      <div class="saas-card"><div class="saas-card-label">Acima de 60 dias</div>
        <div class="saas-card-value c-red">{over60}</div><div class="saas-card-sub">prioridade</div></div>
      <div class="saas-card"><div class="saas-card-label">Acima de 100 dias</div>
        <div class="saas-card-value c-red">{over100}</div><div class="saas-card-sub">crítico</div></div>
      <div class="saas-card"><div class="saas-card-label">Provedores</div>
        <div class="saas-card-value c-teal">{providers}</div><div class="saas-card-sub">afetados</div></div>
      <div class="saas-card"><div class="saas-card-label">Abertos hoje</div>
        <div class="saas-card-value c-purple">{opened}</div><div class="saas-card-sub">na data do relatório</div></div>
    </div>
    """.format(
        total=indicators.get("total_open", 0),
        over30=indicators.get("over_30", 0),
        over60=indicators.get("over_60", 0),
        over100=indicators.get("over_100", 0),
        providers=indicators.get("affected_providers", 0),
        opened=indicators.get("opened_today", 0),
    ), unsafe_allow_html=True)

    st.markdown("#### 1. Resumo das Atividades do Dia")
    report_date = closing.get("report_date_display", closing.get("report_date", ""))
    st.write(
        f"No dia {report_date}, foram realizadas atividades de acompanhamento, atualização e tratamento "
        "dos chamados de Rede Externa vinculados ao projeto."
    )
    st.markdown("**Chamados/INEPs atualizados no sistema:**")
    _render_item_list(closing.get("updated_items", []), "Não foram informados chamados atualizados para esta data.")
    st.markdown("**Chamados/INEPs encerrados:**")
    _render_item_list(closing.get("closed_items", []), "Não foram informados chamados encerrados para esta data.")
    st.write(
        "Além disso, foi mantido contato e monitoramento junto às provedoras responsáveis. "
        "Os casos pendentes permanecem aguardando atuação das operadoras, com prioridade para chamados "
        "classificados como urgentes e com maior tempo em aberto."
    )

    st.markdown("#### 2. Atividades Executadas Hoje")
    for item in [
        "Atualização de chamados no portal;",
        "Validação de status das escolas;",
        "Consulta de provedores responsáveis;",
        "Priorização dos chamados críticos;",
        "Acompanhamento dos casos com mais de 30/60/100 dias;",
        "Registro de observações operacionais;",
        "Encaminhamento/monitoramento junto às provedoras.",
    ]:
        st.markdown(f"- {item}")
    if closing.get("observations"):
        st.markdown("**Complemento operacional:**")
        st.write(closing["observations"])

    st.markdown("#### 3. Indicadores Gerais")
    st.dataframe(_indicators_dataframe(indicators), use_container_width=True, hide_index=True)

    st.markdown("#### 4. Chamados Mais Criticos")
    critical_df = pd.DataFrame(closing.get("critical_rows", []))
    if critical_df.empty:
        st.info("Não há chamados críticos para exibir.")
    else:
        columns = [c for c in ["Ticket/OS", "INEP", "Escola", "UF", "Provedor", "Dias em aberto", "Analista"] if c in critical_df.columns]
        st.dataframe(critical_df[columns], use_container_width=True, hide_index=True, height=420)

    st.markdown("#### 5. Chamados por Provedor")
    provider_df = pd.DataFrame(closing.get("provider_summary", []))
    if provider_df.empty:
        st.info("não h? dados por provedor.")
    else:
        st.dataframe(provider_df, use_container_width=True, hide_index=True)

    st.markdown("#### 6. Distribuição por Estado")
    uf_df = pd.DataFrame(closing.get("uf_summary", []))
    if uf_df.empty:
        st.info("não h? dados por UF.")
    else:
        st.dataframe(uf_df, use_container_width=True, hide_index=True)

    st.markdown("#### 7. Chamados Abertos Hoje")
    opened_df = pd.DataFrame(closing.get("opened_today_rows", []))
    if opened_df.empty:
        if closing.get("has_opening_date_column"):
            st.info("A coluna de abertura foi encontrada, mas nenhum chamado tem data igual à data do relatório.")
        else:
            st.info("A planilha não contém uma coluna de data de abertura reconhecida.")
    else:
        columns = [c for c in ["Ticket/OS", "INEP", "Escola", "UF", "Provedor", "Data de abertura"] if c in opened_df.columns]
        st.dataframe(opened_df[columns], use_container_width=True, hide_index=True)

    st.markdown("#### 8. Todos os Chamados de Rede Externa")
    full_df = pd.DataFrame(closing.get("full_rows", []))
    if full_df.empty:
        st.info("A listagem completa não est? dispon?vel para este fechamento.")
    else:
        col_search, col_uf, col_provider = st.columns([2, 1, 1])
        query = col_search.text_input("Pesquisar Ticket, INEP ou Escola", key="daily_full_search")
        uf_options = sorted([x for x in full_df.get("UF", pd.Series(dtype=str)).dropna().astype(str).unique() if x])
        provider_options = sorted([x for x in full_df.get("Provedor", pd.Series(dtype=str)).dropna().astype(str).unique() if x])
        uf_filter = col_uf.multiselect("UF", uf_options, key="daily_full_uf")
        provider_filter = col_provider.multiselect("Provedor", provider_options, key="daily_full_provider")

        filtered = full_df.copy()
        if query.strip():
            q = query.strip()
            search_cols = [c for c in ["Ticket/OS", "INEP", "Escola"] if c in filtered.columns]
            mask = filtered[search_cols].astype(str).apply(
                lambda row: row.str.contains(q, case=False, na=False).any(), axis=1
            )
            filtered = filtered[mask]
        if uf_filter:
            filtered = filtered[filtered["UF"].isin(uf_filter)]
        if provider_filter:
            filtered = filtered[filtered["Provedor"].isin(provider_filter)]
        display_cols = [c for c in ["Ticket/OS", "Provedor", "Dias em aberto", "INEP", "Escola"] if c in filtered.columns]
        st.dataframe(filtered[display_cols], use_container_width=True, hide_index=True, height=520)

    st.markdown("#### 9. Pontos de Atenção")
    all_alerts = closing.get("alerts", []) + closing.get("manual_attention_items", [])
    _render_item_list(all_alerts, "não h? pontos de Atenção adicionais registrados.")

    st.markdown("#### 10. Observações Operacionais")
    st.write(closing.get("observations") or "não foram registradas Observações operacionais adicionais.")

    st.markdown("#### 11. Proximas atividades")
    st.write(closing.get("next_actions") or "Manter acompanhamento dos chamados críticos e pendências com as provedoras.")

    export_pdf_col, export_excel_col = st.columns(2)
    pdf_bytes = generate_daily_report_pdf(closing)
    excel_bytes = generate_daily_report_excel(closing)
    export_pdf_col.download_button(
        "Baixar relat\u00f3rio em PDF",
        pdf_bytes,
        file_name=f"relatorio_diario_{closing.get('report_date', date.today().isoformat())}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )
    export_excel_col.download_button(
        "Baixar relat\u00f3rio em Excel",
        excel_bytes,
        file_name=f"relatorio_monitoramento_{closing.get('report_date', date.today().isoformat())}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


def render_daily_closing_admin() -> None:
    st.markdown('<div class="sidebar-label">Área Restrita</div>', unsafe_allow_html=True)
    with st.expander("Fechamento Diário", expanded=False):
        if not st.session_state.get("daily_report_authenticated"):
            password = st.text_input("Senha", type="password", key="daily_report_password_input")
            if st.button("Entrar", key="daily_report_login", use_container_width=True):
                if check_daily_report_password(password):
                    st.session_state.daily_report_authenticated = True
                    st.success("Acesso liberado.")
                    st.rerun()
                else:
                    st.error("Senha inválida.")
            return

        st.caption("Upload e edição ficam restritos a esta área.")
        report_date = st.date_input("Data do relatório", value=date.today(), key="daily_report_date")
        responsible = st.text_input("Responsável", key="daily_report_responsible")
        uploaded = st.file_uploader(
            "Planilha diária de chamados (.xlsx)",
            type=["xlsx"],
            key="daily_report_upload",
        )
        updated_text = st.text_area("Chamados/INEPs atualizados hoje", height=110, key="daily_report_updated")
        closed_text = st.text_area("Chamados/INEPs encerrados hoje", height=110, key="daily_report_closed")
        observations = st.text_area("Observações operacionais", height=120, key="daily_report_observations")
        attention = st.text_area("Pontos de atenção", height=100, key="daily_report_attention")
        next_actions = st.text_area("Próximas ações recomendadas", height=100, key="daily_report_next_actions")

        if st.button("Concluir Fechamento", type="primary", use_container_width=True):
            closing, error = build_daily_closing(
                report_date=report_date,
                responsible=responsible,
                uploaded_file=uploaded,
                updated_text=updated_text,
                closed_text=closed_text,
                observations=observations,
                attention_points=attention,
                next_actions=next_actions,
            )
            if error:
                st.error(error)
            else:
                st.session_state.daily_report_current = closing
                save_daily_closing(closing)
                st.success("Fechamento concluído e relatório atualizado.")
                st.rerun()

        if st.button("Limpar fechamento atual", use_container_width=True):
            st.session_state.daily_report_current = None
            clear_current_daily_closing()
            st.success("Fechamento atual limpo. O histórico foi mantido para consulta.")
            st.rerun()

        if st.button("Sair da área restrita", use_container_width=True):
            st.session_state.daily_report_authenticated = False
            st.rerun()
