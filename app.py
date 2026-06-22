import streamlit as st
import requests
import pandas as pd
import json
import hmac
import hashlib
import html
import os
import re
import secrets as token_secrets
import smtplib
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from urllib.parse import quote, urlsplit
from zoneinfo import ZoneInfo

from modules.daily_report import render_daily_closing_admin, render_daily_report

# ═══════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════
CONFIG_FILE       = "config_contas.json"
CONFIG_PROVEDORES = "config_provedores.json"
BASE_URL          = "https://api.ui.com/v1"
APP_TZ            = ZoneInfo("America/Sao_Paulo")
LAST_SEEN_COLUMN  = "Último Sinal"
HUB_REGISTRATION_EMAIL = "breno.brocanello@gmail.com"
HUB_ACCESS_REQUESTS_FILE = os.path.join("data", "solicitacoes_acesso.json")
HUB_PASSWORD_HASH_ITERATIONS = 260_000
HUB_ALLOWED_EMAILS = set()

st.set_page_config(
    page_title="Hub Redes — EACE",
    page_icon="🛰️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ═══════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

:root, .stApp {
    --hub-bg: var(--background-color, #f3f6fb);
    --hub-sidebar-bg: var(--secondary-background-color, #ffffff);
    --hub-surface: var(--secondary-background-color, #ffffff);
    --hub-surface-2: var(--background-color, #f8fafc);
    --hub-border: #d7dee9;
    --hub-text: var(--text-color, #111827);
    --hub-muted: color-mix(in srgb, var(--hub-text) 72%, transparent);
    --hub-muted-2: color-mix(in srgb, var(--hub-text) 58%, transparent);
    --hub-log-bg: var(--secondary-background-color, #f8fafc);
    --hub-scroll: #cbd5e1;
    --hub-accent: #ef4444;
    --hub-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    --hub-blue: #2563eb;
    --hub-green: #16a34a;
    --hub-red: #dc2626;
    --hub-yellow: #ca8a04;
    --hub-gray: #6b7280;
    --hub-teal: #0891b2;
    --hub-purple: #7c3aed;
}

@media (prefers-color-scheme: dark) {
    :root, .stApp {
        --hub-bg: #0f1117;
        --hub-sidebar-bg: #161b27;
        --hub-surface: #161b27;
        --hub-surface-2: #0a0d14;
        --hub-border: #273244;
        --hub-text: #f8fafc;
        --hub-muted: #cbd5e1;
        --hub-muted-2: #94a3b8;
        --hub-log-bg: #0a0d14;
        --hub-scroll: #475569;
        --hub-accent: #f87171;
        --hub-shadow: none;
        --hub-blue: #60a5fa;
        --hub-green: #4ade80;
        --hub-red: #f87171;
        --hub-yellow: #facc15;
        --hub-gray: #94a3b8;
        --hub-teal: #22d3ee;
        --hub-purple: #c084fc;
    }
}

html[data-theme="light"], body[data-theme="light"], .stApp[data-theme="light"],
[data-testid="stAppViewContainer"][data-theme="light"], [data-baseweb-theme="light"],
[data-theme="light"] {
    --hub-bg: #f3f6fb;
    --hub-sidebar-bg: #ffffff;
    --hub-surface: #ffffff;
    --hub-surface-2: #f8fafc;
    --hub-border: #d7dee9;
    --hub-text: #111827;
    --hub-muted: #475569;
    --hub-muted-2: #64748b;
    --hub-log-bg: #f8fafc;
    --hub-scroll: #cbd5e1;
    --hub-accent: #ef4444;
    --hub-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    --hub-blue: #2563eb;
    --hub-green: #16a34a;
    --hub-red: #dc2626;
    --hub-yellow: #ca8a04;
    --hub-gray: #6b7280;
    --hub-teal: #0891b2;
    --hub-purple: #7c3aed;
}

html[data-theme="dark"], body[data-theme="dark"], .stApp[data-theme="dark"],
[data-testid="stAppViewContainer"][data-theme="dark"], [data-baseweb-theme="dark"],
[data-theme="dark"], .dark {
    --hub-bg: #0f1117;
    --hub-sidebar-bg: #161b27;
    --hub-surface: #161b27;
    --hub-surface-2: #0a0d14;
    --hub-border: #273244;
    --hub-text: #f8fafc;
    --hub-muted: #cbd5e1;
    --hub-muted-2: #94a3b8;
    --hub-log-bg: #0a0d14;
    --hub-scroll: #475569;
    --hub-accent: #f87171;
    --hub-shadow: none;
    --hub-blue: #60a5fa;
    --hub-green: #4ade80;
    --hub-red: #f87171;
    --hub-yellow: #facc15;
    --hub-gray: #94a3b8;
    --hub-teal: #22d3ee;
    --hub-purple: #c084fc;
}

html, body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background: var(--hub-bg) !important;
    color: var(--hub-text) !important;
}

[data-testid="stHeader"] {
    background: var(--hub-bg) !important;
    box-shadow: none !important;
}

[data-testid="stSidebar"] > div:first-child {
    background: var(--hub-sidebar-bg) !important;
    border-right: 1px solid var(--hub-border) !important;
}

.block-container { padding-top: 3.4rem !important; max-width: 1400px !important; }
* { word-break: break-word !important; overflow-wrap: break-word !important; }

.saas-grid {
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    gap: 10px;
    margin: 20px 0 24px 0;
}
.saas-grid-4 {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
    margin: 20px 0 24px 0;
}
.saas-card {
    background: var(--hub-surface);
    border: 1px solid var(--hub-border);
    border-radius: 10px;
    padding: 16px 18px;
    min-width: 0;
    box-shadow: var(--hub-shadow);
}
.saas-card-label {
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    color: var(--hub-muted);
    margin-bottom: 8px;
    font-family: Inter, sans-serif;
}
.saas-card-value {
    font-size: 28px;
    font-weight: 700;
    font-family: Inter, sans-serif;
    line-height: 1;
}
.saas-card-sub {
    font-size: 11px;
    color: var(--hub-muted-2);
    margin-top: 6px;
    font-family: Inter, sans-serif;
}
.c-blue   { color: var(--hub-blue); }
.c-green  { color: var(--hub-green); }
.c-red    { color: var(--hub-red); }
.c-yellow { color: var(--hub-yellow); }
.c-gray   { color: var(--hub-gray); }
.c-teal   { color: var(--hub-teal); }
.c-purple { color: var(--hub-purple); }

.page-title {
    font-family: Inter, sans-serif;
    font-size: 22px;
    font-weight: 700;
    color: var(--hub-text);
    margin: 0;
    line-height: 1.35;
    min-height: 32px;
    padding-top: 2px;
    overflow: visible;
}
.page-sub {
    font-family: Inter, sans-serif;
    font-size: 12px;
    color: var(--hub-muted);
    margin-top: 3px;
    letter-spacing: 0.5px;
    line-height: 1.35;
}

.sidebar-label {
    font-family: Inter, sans-serif;
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: var(--hub-muted-2);
    margin: 16px 0 8px 0;
    padding-bottom: 6px;
    border-bottom: 1px solid var(--hub-border);
}

.log-box {
    background: var(--hub-log-bg);
    border: 1px solid var(--hub-border);
    border-radius: 8px;
    padding: 12px 16px;
    font-family: 'Courier New', monospace;
    font-size: 12px;
    line-height: 1.8;
    max-height: 160px;
    overflow-y: auto;
    margin: 12px 0;
}
.log-ok   { color: #22c55e; }
.log-err  { color: #ef4444; }
.log-info { color: #3b82f6; }

.ts-pill {
    display: inline-block;
    background: var(--hub-surface);
    border: 1px solid var(--hub-border);
    border-radius: 6px;
    padding: 3px 10px;
    font-size: 11px;
    color: #3b82f6;
    font-family: Inter, sans-serif;
    margin-bottom: 16px;
}

.saas-divider {
    border: none;
    border-top: 1px solid var(--hub-border);
    margin: 20px 0;
}

::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--hub-scroll); border-radius: 99px; }

[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"],
[data-testid="stSidebar"] [data-testid="stMetricLabel"],
[data-testid="stSidebar"] [data-testid="stMetricValue"] {
    color: var(--hub-text) !important;
}

[data-testid="stSidebar"] .sidebar-label {
    color: var(--hub-muted) !important;
}

[data-testid="stSidebar"] [data-baseweb="tag"] span {
    color: #ffffff !important;
}

[data-testid="stTabs"] button {
    color: var(--hub-muted) !important;
}

[data-testid="stTabs"] button p {
    color: var(--hub-muted) !important;
    font-weight: 600;
}

[data-testid="stTabs"] button[aria-selected="true"] p {
    color: var(--hub-accent) !important;
}

[data-testid="stTabs"] [data-baseweb="tab-highlight"] {
    background-color: var(--hub-accent) !important;
}

h1, h2, h3, h4, h5, h6,
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li {
    color: var(--hub-text);
}

[data-testid="stAlert"] {
    border-radius: 8px;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# PERSISTÊNCIA
# ═══════════════════════════════════════════════════════════════
def salvar_contas(contas):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(contas, f, indent=2)
    except Exception:
        pass

# Prefixo usado para identificar chaves de conta nos Secrets raiz
_CONTA_PREFIXOS = ("Conta_", "conta_", "CONTA_", "admin")

def carregar_contas():
    """
    Lê as contas dos Streamlit Secrets.
    Suporta quatro formatos:

    Formato 1 — seção [contas] (ideal):
        [contas]
        Conta_AC1 = "chave1"

    Formato 2 — chaves na raiz com prefixo Conta_:
        Conta_AC1 = "chave1"
        admin0    = "chave2"

    Formato 3 — JSON na chave contas_json:
        contas_json = '[{"apelido":"Conta AC1","api_key":"chave1"}]'

    Formato 4 — fallback arquivo local (desenvolvimento).
    """
    # Formato 1: seção [contas]
    try:
        if "contas" in st.secrets:
            sec = st.secrets["contas"]
            resultado = [
                {"apelido": str(k), "api_key": str(v)}
                for k, v in sec.items()
                if str(v).strip()
            ]
            if resultado:
                return resultado
    except Exception:
        pass

    # Formato 2: chaves soltas na raiz que parecem contas Ubiquiti
    # O Streamlit Cloud às vezes não processa seções TOML corretamente
    try:
        todas = dict(st.secrets)
        contas_raiz = []
        chaves_ignorar = {"contas_json"}
        for k, v in todas.items():
            if k in chaves_ignorar:
                continue
            v_str = str(v).strip()
            # Considera conta se: começa com prefixo conhecido OU o valor parece uma API key
            # (string sem espaços com 20+ chars)
            eh_prefixo = any(k.startswith(p) for p in _CONTA_PREFIXOS)
            eh_apikey  = len(v_str) >= 20 and " " not in v_str and isinstance(v, str)
            if eh_prefixo or eh_apikey:
                contas_raiz.append({"apelido": k, "api_key": v_str})
        if contas_raiz:
            return contas_raiz
    except Exception:
        pass

    # Formato 3: JSON na chave contas_json
    try:
        if "contas_json" in st.secrets:
            return json.loads(st.secrets["contas_json"])
    except Exception:
        pass

    # Formato 4: fallback arquivo local
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def salvar_provedores(provs: dict) -> bool:
    try:
        with open(CONFIG_PROVEDORES, "w", encoding="utf-8") as f:
            json.dump(provs, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        st.error(f"Erro ao salvar provedores: {e}")
        return False

def carregar_provedores() -> dict:
    if os.path.exists(CONFIG_PROVEDORES):
        try:
            with open(CONFIG_PROVEDORES, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
        except Exception:
            pass
    return {}


def agora_iso() -> str:
    return datetime.now(APP_TZ).isoformat(timespec="seconds")


def obter_config(*chaves, default: str = "") -> str:
    for chave in chaves:
        valor = None
        try:
            if isinstance(chave, tuple):
                atual = st.secrets
                for parte in chave:
                    atual = atual[parte]
                valor = atual
            elif chave in st.secrets:
                valor = st.secrets[chave]
        except Exception:
            valor = None

        if valor is None and isinstance(chave, str):
            valor = os.getenv(chave) or os.getenv(chave.upper())

        if valor is not None and str(valor).strip():
            return str(valor).strip()
    return default


def obter_url_publica_hub() -> str:
    url = obter_config("hub_public_url", "HUB_PUBLIC_URL", default="").rstrip("/")
    if url:
        return url

    try:
        current_url = str(getattr(st.context, "url", "") or "")
        if current_url:
            parsed = urlsplit(current_url)
            if parsed.scheme and parsed.netloc:
                base_path = parsed.path.rstrip("/")
                return f"{parsed.scheme}://{parsed.netloc}{base_path}"
    except Exception:
        pass

    return "http://localhost:8501"


def obter_config_supabase() -> tuple[str, str, str, str]:
    url = obter_config(("supabase", "url"), "supabase_url", "SUPABASE_URL", default="").rstrip("/")
    key = obter_config(
        ("supabase", "secret_key"),
        ("supabase", "service_role_key"),
        "supabase_secret_key",
        "SUPABASE_SECRET_KEY",
        "supabase_service_role_key",
        "SUPABASE_SERVICE_ROLE_KEY",
        default="",
    )
    requests_table = obter_config(
        ("supabase", "access_requests_table"),
        "supabase_access_requests_table",
        default="hub_access_requests",
    )
    users_table = obter_config(
        ("supabase", "allowed_users_table"),
        "supabase_allowed_users_table",
        default="hub_allowed_users",
    )
    return url, key, requests_table, users_table


def supabase_ativo() -> bool:
    url, key, _, _ = obter_config_supabase()
    return bool(url and key)


def supabase_headers(extra: dict | None = None) -> dict:
    _, key, _, _ = obter_config_supabase()
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if extra:
        headers.update(extra)
    return headers


def supabase_request(method: str, table: str, query: str = "", payload=None, prefer: str | None = None):
    url, _, _, _ = obter_config_supabase()
    endpoint = f"{url}/rest/v1/{table}{query}"
    extra = {"Prefer": prefer} if prefer else None
    resp = requests.request(
        method,
        endpoint,
        headers=supabase_headers(extra),
        json=payload,
        timeout=15,
    )
    if resp.status_code >= 400:
        detalhe = resp.text[:300] if resp.text else resp.reason
        raise RuntimeError(f"Supabase retornou {resp.status_code}: {detalhe}")
    if not resp.text:
        return None
    try:
        return resp.json()
    except ValueError:
        return None


def carregar_solicitacoes_locais() -> list[dict]:
    if not os.path.exists(HUB_ACCESS_REQUESTS_FILE):
        return []
    try:
        with open(HUB_ACCESS_REQUESTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def salvar_solicitacoes_locais(solicitacoes: list[dict]) -> None:
    os.makedirs(os.path.dirname(HUB_ACCESS_REQUESTS_FILE), exist_ok=True)
    with open(HUB_ACCESS_REQUESTS_FILE, "w", encoding="utf-8") as f:
        json.dump(solicitacoes, f, indent=2, ensure_ascii=False)


def normalizar_email(email: str) -> str:
    return str(email or "").strip().lower()


def email_parece_valido(email: str) -> bool:
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", normalizar_email(email)))


def hash_token(token: str) -> str:
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def gerar_hash_senha(senha: str) -> str:
    salt = token_secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        str(senha or "").encode("utf-8"),
        salt.encode("utf-8"),
        HUB_PASSWORD_HASH_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${HUB_PASSWORD_HASH_ITERATIONS}${salt}${digest}"


def verificar_hash_senha(senha: str, senha_hash: str) -> bool:
    partes = str(senha_hash or "").split("$")
    if len(partes) != 4 or partes[0] != "pbkdf2_sha256":
        return False
    try:
        iteracoes = int(partes[1])
        salt = partes[2]
        digest_salvo = partes[3]
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            str(senha or "").encode("utf-8"),
            salt.encode("utf-8"),
            iteracoes,
        ).hex()
        return hmac.compare_digest(digest, digest_salvo)
    except Exception:
        return False


def validar_senha_registro(senha: str, confirmacao: str) -> tuple[bool, str]:
    if not senha:
        return False, "Informe uma senha."
    if len(senha) < 8:
        return False, "A senha precisa ter pelo menos 8 caracteres."
    if senha != confirmacao:
        return False, "A confirmação da senha não confere."
    return True, ""


def carregar_usuarios_aprovados_supabase() -> dict[str, dict]:
    if not supabase_ativo():
        return {}
    _, _, _, users_table = obter_config_supabase()
    try:
        rows = supabase_request(
            "GET",
            users_table,
            "?select=email,name,password_hash&active=eq.true",
        ) or []
        return {
            normalizar_email(row.get("email")): row
            for row in rows
            if row.get("email")
        }
    except Exception:
        return {}


def carregar_usuarios_aprovados_locais() -> dict[str, dict]:
    usuarios = {}
    for item in carregar_solicitacoes_locais():
        if item.get("status") == "approved" and item.get("email"):
            usuarios[normalizar_email(item["email"])] = item
    return usuarios


def carregar_usuarios_aprovados() -> dict[str, dict]:
    usuarios = carregar_usuarios_aprovados_locais()
    usuarios.update(carregar_usuarios_aprovados_supabase())
    return usuarios


def carregar_emails_permitidos() -> set[str]:
    emails = {normalizar_email(email) for email in HUB_ALLOWED_EMAILS}
    emails.update(carregar_usuarios_aprovados().keys())
    return emails


def credenciais_hub_validas(email: str, senha: str) -> bool:
    email_normalizado = normalizar_email(email)
    usuario = carregar_usuarios_aprovados().get(email_normalizado)
    if not usuario:
        return False
    return verificar_hash_senha(senha, usuario.get("password_hash", ""))


def montar_links_decisao_solicitacao(request_id: str, token: str) -> tuple[str, str]:
    base_url = obter_url_publica_hub()
    approve_url = f"{base_url}?access_action=approve&request_id={quote(request_id)}&approval_token={quote(token)}"
    reject_url = f"{base_url}?access_action=reject&request_id={quote(request_id)}&approval_token={quote(token)}"
    return approve_url, reject_url


def enviar_email_smtp(to_email: str, subject: str, text: str, html_body: str) -> tuple[bool, str]:
    host = obter_config(("smtp", "host"), "smtp_host", "SMTP_HOST", default="smtp.gmail.com")
    user = obter_config(
        ("smtp", "user"),
        "smtp_user",
        "SMTP_USER",
        "gmail_user",
        "GMAIL_USER",
        default=HUB_REGISTRATION_EMAIL,
    )
    smtp_password = obter_config(("smtp", "password"), "smtp_password", "SMTP_PASSWORD", default="")
    gmail_password = obter_config("gmail_app_password", "GMAIL_APP_PASSWORD", default="").replace(" ", "")
    password = smtp_password or gmail_password
    if not password:
        return False, "Envio de e-mail não configurado: defina gmail_app_password nos Secrets."

    port = int(obter_config(("smtp", "port"), "smtp_port", "SMTP_PORT", default="587"))
    from_email = obter_config(("smtp", "from_email"), "smtp_from_email", "SMTP_FROM_EMAIL", default=user)

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email
    msg.attach(MIMEText(text, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    try:
        with smtplib.SMTP(host, port, timeout=15) as server:
            server.starttls()
            server.login(user, password)
            server.sendmail(from_email, [to_email], msg.as_string())
        return True, "E-mail enviado por SMTP."
    except Exception as exc:
        return False, f"Falha no SMTP: {exc}"


def enviar_email_resend(to_email: str, subject: str, text: str, html_body: str) -> tuple[bool, str]:
    api_key = obter_config(("resend", "api_key"), "resend_api_key", "RESEND_API_KEY", default="")
    if not api_key:
        return False, "Resend não configurado."

    from_email = obter_config(
        ("resend", "from_email"),
        "resend_from_email",
        "RESEND_FROM_EMAIL",
        default="Hub Ubiquiti <onboarding@resend.dev>",
    )
    payload = {
        "from": from_email,
        "to": [to_email],
        "subject": subject,
        "text": text,
        "html": html_body,
    }
    try:
        resp = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=15,
        )
        if resp.status_code >= 400:
            return False, f"Resend retornou {resp.status_code}: {resp.text[:300]}"
        return True, "E-mail enviado por Resend."
    except Exception as exc:
        return False, f"Falha no Resend: {exc}"


def enviar_email_solicitacao_acesso(solicitacao: dict, approve_url: str, reject_url: str) -> tuple[bool, str]:
    nome = html.escape(solicitacao.get("name") or "Não informado")
    email = html.escape(solicitacao.get("email") or "")
    motivo = html.escape(solicitacao.get("reason") or "Não informado")
    criado_em = html.escape(solicitacao.get("created_at") or agora_iso())
    subject = f"Solicitação de acesso ao Hub Ubiquiti - {email}"
    text = (
        "Nova solicitação de acesso ao Hub Ubiquiti\n\n"
        f"Nome: {solicitacao.get('name') or 'Não informado'}\n"
        f"E-mail: {solicitacao.get('email')}\n"
        f"Motivo: {solicitacao.get('reason') or 'Não informado'}\n"
        f"Criado em: {solicitacao.get('created_at') or agora_iso()}\n\n"
        f"Aprovar: {approve_url}\n"
        f"Recusar: {reject_url}\n"
    )
    html_body = f"""
    <div style="font-family: Arial, sans-serif; color: #111827; line-height: 1.5;">
        <h2>Nova solicitação de acesso</h2>
        <p><strong>Nome:</strong> {nome}</p>
        <p><strong>E-mail:</strong> {email}</p>
        <p><strong>Motivo:</strong> {motivo}</p>
        <p><strong>Criado em:</strong> {criado_em}</p>
        <p style="margin-top: 24px;">
            <a href="{html.escape(approve_url)}"
               style="background:#16a34a;color:#fff;padding:10px 16px;text-decoration:none;border-radius:6px;margin-right:8px;">
               Aprovar acesso
            </a>
            <a href="{html.escape(reject_url)}"
               style="background:#dc2626;color:#fff;padding:10px 16px;text-decoration:none;border-radius:6px;">
               Recusar
            </a>
        </p>
    </div>
    """

    ok, msg = enviar_email_resend(HUB_REGISTRATION_EMAIL, subject, text, html_body)
    if ok:
        return ok, msg
    return enviar_email_smtp(HUB_REGISTRATION_EMAIL, subject, text, html_body)


def salvar_solicitacao_supabase(solicitacao: dict) -> None:
    _, _, requests_table, _ = obter_config_supabase()
    supabase_request(
        "POST",
        requests_table,
        "?on_conflict=id",
        payload=solicitacao,
        prefer="resolution=merge-duplicates,return=minimal",
    )


def salvar_usuario_aprovado_supabase(solicitacao: dict) -> None:
    _, _, _, users_table = obter_config_supabase()
    supabase_request(
        "POST",
        users_table,
        "?on_conflict=email",
        payload={
            "email": normalizar_email(solicitacao["email"]),
            "name": solicitacao.get("name") or "",
            "password_hash": solicitacao.get("password_hash") or "",
            "active": True,
            "approved_at": solicitacao.get("decided_at") or agora_iso(),
            "approved_by": solicitacao.get("decided_by") or HUB_REGISTRATION_EMAIL,
        },
        prefer="resolution=merge-duplicates,return=minimal",
    )


def migrar_solicitacoes_locais_para_supabase() -> tuple[int, int]:
    if not supabase_ativo():
        return 0, 0

    migradas = 0
    aprovadas = 0
    for solicitacao in carregar_solicitacoes_locais():
        if not solicitacao.get("id") or not solicitacao.get("email"):
            continue
        salvar_solicitacao_supabase(solicitacao)
        migradas += 1
        if solicitacao.get("status") == "approved":
            salvar_usuario_aprovado_supabase(solicitacao)
            aprovadas += 1
    return migradas, aprovadas


def salvar_solicitacao_local(solicitacao: dict) -> None:
    solicitacoes = carregar_solicitacoes_locais()
    solicitacoes.append(solicitacao)
    salvar_solicitacoes_locais(solicitacoes)


def criar_solicitacao_acesso(nome: str, email: str, motivo: str, senha: str, confirmacao_senha: str) -> tuple[bool, str]:
    email_normalizado = normalizar_email(email)
    if not nome.strip():
        return False, "Informe seu nome."
    if not email_parece_valido(email_normalizado):
        return False, "Informe um e-mail válido."
    senha_ok, senha_msg = validar_senha_registro(senha, confirmacao_senha)
    if not senha_ok:
        return False, senha_msg
    if email_normalizado in carregar_emails_permitidos():
        return False, "Este e-mail já está liberado para acessar o hub."

    request_id = token_secrets.token_hex(12)
    token = token_secrets.token_urlsafe(32)
    solicitacao = {
        "id": request_id,
        "name": nome.strip(),
        "email": email_normalizado,
        "reason": motivo.strip(),
        "status": "pending",
        "token_hash": hash_token(token),
        "password_hash": gerar_hash_senha(senha),
        "created_at": agora_iso(),
        "decided_at": None,
        "decided_by": None,
    }

    try:
        if supabase_ativo():
            salvar_solicitacao_supabase(solicitacao)
        else:
            salvar_solicitacao_local(solicitacao)
    except Exception as exc:
        return False, f"Não foi possível registrar a solicitação: {exc}"

    approve_url, reject_url = montar_links_decisao_solicitacao(request_id, token)
    email_ok, email_msg = enviar_email_solicitacao_acesso(solicitacao, approve_url, reject_url)
    if not email_ok:
        return True, f"Solicitação registrada, mas o e-mail ao administrador não foi enviado: {email_msg}"
    return True, "Solicitação enviada. O acesso ficará pendente até aprovação."


def buscar_solicitacao_supabase(request_id: str) -> dict | None:
    _, _, requests_table, _ = obter_config_supabase()
    rows = supabase_request(
        "GET",
        requests_table,
        f"?select=*&id=eq.{quote(request_id)}&limit=1",
    ) or []
    return rows[0] if rows else None


def atualizar_solicitacao_supabase(solicitacao: dict, status: str) -> None:
    _, _, requests_table, _ = obter_config_supabase()
    payload = {
        "status": status,
        "decided_at": agora_iso(),
        "decided_by": HUB_REGISTRATION_EMAIL,
    }
    supabase_request(
        "PATCH",
        requests_table,
        f"?id=eq.{quote(solicitacao['id'])}",
        payload=payload,
        prefer="return=minimal",
    )
    if status == "approved":
        solicitacao_aprovada = dict(solicitacao)
        solicitacao_aprovada.update(payload)
        salvar_usuario_aprovado_supabase(solicitacao_aprovada)


def buscar_solicitacao_local(request_id: str) -> dict | None:
    for solicitacao in carregar_solicitacoes_locais():
        if solicitacao.get("id") == request_id:
            return solicitacao
    return None


def atualizar_solicitacao_local(request_id: str, status: str) -> None:
    solicitacoes = carregar_solicitacoes_locais()
    for solicitacao in solicitacoes:
        if solicitacao.get("id") == request_id:
            solicitacao["status"] = status
            solicitacao["decided_at"] = agora_iso()
            solicitacao["decided_by"] = HUB_REGISTRATION_EMAIL
            break
    salvar_solicitacoes_locais(solicitacoes)


def decidir_solicitacao_acesso(request_id: str, token: str, action: str) -> tuple[bool, str]:
    if action not in {"approve", "reject"}:
        return False, "Ação inválida."
    if not request_id or not token:
        return False, "Link de decisão inválido."

    origem_local = False
    try:
        if supabase_ativo():
            solicitacao = buscar_solicitacao_supabase(request_id)
            if not solicitacao:
                solicitacao = buscar_solicitacao_local(request_id)
                origem_local = bool(solicitacao)
        else:
            solicitacao = buscar_solicitacao_local(request_id)
    except Exception as exc:
        return False, f"Não foi possível consultar a solicitação: {exc}"

    if not solicitacao:
        return False, "Solicitação não encontrada."
    if not hmac.compare_digest(str(solicitacao.get("token_hash", "")), hash_token(token)):
        return False, "Token de aprovação inválido."
    if solicitacao.get("status") in {"approved", "rejected"}:
        status_txt = "aprovada" if solicitacao.get("status") == "approved" else "recusada"
        return True, f"Esta solicitação já tinha sido {status_txt}."

    novo_status = "approved" if action == "approve" else "rejected"
    try:
        if supabase_ativo():
            if origem_local:
                salvar_solicitacao_supabase(solicitacao)
            atualizar_solicitacao_supabase(solicitacao, novo_status)
            if origem_local:
                atualizar_solicitacao_local(request_id, novo_status)
        else:
            atualizar_solicitacao_local(request_id, novo_status)
    except Exception as exc:
        return False, f"Não foi possível atualizar a solicitação: {exc}"

    if novo_status == "approved":
        return True, f"Acesso aprovado para {normalizar_email(solicitacao.get('email'))}."
    return True, f"Solicitação recusada para {normalizar_email(solicitacao.get('email'))}."


def obter_query_param(nome: str) -> str:
    try:
        valor = st.query_params.get(nome, "")
        if isinstance(valor, list):
            return str(valor[0]) if valor else ""
        return str(valor)
    except Exception:
        return ""


def processar_link_decisao_acesso() -> bool:
    action = obter_query_param("access_action")
    if action not in {"approve", "reject"}:
        return False

    request_id = obter_query_param("request_id")
    token = obter_query_param("approval_token")
    ok, msg = decidir_solicitacao_acesso(request_id, token, action)

    st.markdown(
        """
        <div style="max-width: 540px; margin: 10vh auto 24px auto; text-align: center;">
            <p class="page-title" style="font-size: 30px; margin-bottom: 8px;">Solicitação de acesso</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if ok:
        st.success(msg)
    else:
        st.error(msg)
    if st.button("Voltar para o login", type="primary"):
        try:
            st.query_params.clear()
        except Exception:
            pass
        st.rerun()
    return True


def render_hub_login() -> None:
    st.markdown(
        """
        <div style="max-width: 460px; margin: 7vh auto 22px auto; text-align: center;">
            <p class="page-title" style="font-size: 34px; margin-bottom: 8px;">Hub Redes - EACE</p>
            <p class="page-sub" style="margin: 0;">Acesso restrito ao hub geral</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    _, login_col, _ = st.columns([1, 1.1, 1])
    with login_col:
        login_tab, register_tab = st.tabs(["Entrar", "Registre-se"])

        with login_tab:
            with st.form("hub_login_form"):
                email = st.text_input("E-mail", placeholder="seu.email@gmail.com")
                senha = st.text_input("Senha", type="password")
                entrar = st.form_submit_button("Entrar", type="primary", use_container_width=True)

            if entrar:
                if credenciais_hub_validas(email, senha):
                    st.session_state.hub_authenticated = True
                    st.session_state.hub_user_email = normalizar_email(email)
                    st.rerun()
                else:
                    st.error("E-mail ou senha inválidos.")

        with register_tab:
            with st.form("hub_register_form"):
                nome = st.text_input("Nome completo")
                email_registro = st.text_input("E-mail corporativo ou Gmail", key="hub_register_email")
                senha_registro = st.text_input("Senha", type="password", key="hub_register_password")
                confirmacao_senha = st.text_input(
                    "Confirmar senha",
                    type="password",
                    key="hub_register_password_confirm",
                )
                motivo = st.text_area("Motivo do acesso", height=90)
                registrar = st.form_submit_button("Enviar solicitação", type="primary", use_container_width=True)

            if registrar:
                ok, msg = criar_solicitacao_acesso(
                    nome,
                    email_registro,
                    motivo,
                    senha_registro,
                    confirmacao_senha,
                )
                if ok:
                    if "não foi enviado" in msg:
                        st.warning(msg)
                    else:
                        st.success(msg)
                else:
                    st.error(msg)


# ═══════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════
for k, v in {
    "hub_authenticated": False,
    "hub_user_email": "",
    "hub_local_requests_migrated": False,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

if not st.session_state.hub_local_requests_migrated:
    try:
        migrar_solicitacoes_locais_para_supabase()
    except Exception:
        pass
    st.session_state.hub_local_requests_migrated = True

if processar_link_decisao_acesso():
    st.stop()

if not st.session_state.hub_authenticated:
    render_hub_login()
    st.stop()

for k, v in {
    "contas":            None,
    "provedores":        None,
    "df_resultado":      None,
    "relatorio":         [],
    "ultima_consulta":   None,
    "df_inventario":     None,
    "inv_ultima_coleta": None,
    "daily_report_current": None,
    "daily_report_authenticated": False,
}.items():
    if k not in st.session_state:
        if k == "contas":
            st.session_state[k] = carregar_contas()
        elif k == "provedores":
            st.session_state[k] = carregar_provedores()
        else:
            st.session_state[k] = v

# ═══════════════════════════════════════════════════════════════
# API UBIQUITI
# ═══════════════════════════════════════════════════════════════
def formatar_timestamp_api(valor) -> str:
    if valor in (None, "", "—"):
        return "—"
    try:
        if isinstance(valor, (int, float)):
            ts = float(valor)
            if ts > 10_000_000_000:
                ts = ts / 1000
            dt = datetime.fromtimestamp(ts, tz=timezone.utc)
        else:
            texto = str(valor).strip()
            if not texto:
                return "—"
            if re.fullmatch(r"\d+(\.\d+)?", texto):
                ts = float(texto)
                if ts > 10_000_000_000:
                    ts = ts / 1000
                dt = datetime.fromtimestamp(ts, tz=timezone.utc)
            else:
                dt = datetime.fromisoformat(texto.replace("Z", "+00:00"))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(APP_TZ).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(valor).strip() or "—"

def extrair_ultimo_sinal(host: dict, rep: dict) -> str:
    candidatos = [
        host.get("lastConnectionStateChange"),
        rep.get("deviceStateLastChanged"),
        host.get("latestBackupTime"),
    ]
    grupos = host.get("userData", {}).get("consoleGroupMembers", [])
    if isinstance(grupos, list):
        for grupo in grupos:
            attrs = grupo.get("roleAttributes", {}) if isinstance(grupo, dict) else {}
            candidatos.append(attrs.get("connectedStateLastChanged"))
    for valor in candidatos:
        formatado = formatar_timestamp_api(valor)
        if formatado != "—":
            return formatado
    return "—"

def get_paginated_hosts(api_key: str) -> list:
    items, next_token = [], None
    while True:
        params = {"pageSize": 200}
        if next_token:
            params["nextToken"] = next_token
        r = requests.get(
            f"{BASE_URL}/hosts",
            headers={"X-API-KEY": api_key, "Accept": "application/json"},
            params=params, timeout=25
        )
        if r.status_code == 401: raise Exception("Chave inválida (401)")
        if r.status_code == 429: raise Exception("Rate limit (429) — aguarde 1 min")
        if r.status_code != 200: raise Exception(f"Erro HTTP {r.status_code}")
        data       = r.json()
        batch      = data.get("data", [])
        if not batch: break
        items.extend(batch)
        next_token = data.get("nextToken", "")
        if not next_token: break
        time.sleep(0.3)
    return items

def extrair_host(host: dict) -> dict:
    rep    = host.get("reportedState", {})
    nome   = str(rep.get("name", host.get("name", ""))).strip()
    estado = str(rep.get("state", "disconnected")).lower()
    ip     = str(host.get("ipAddress", rep.get("ip", "—"))).strip()
    ultimo_sinal = extrair_ultimo_sinal(host, rep) if estado != "connected" else "—"
    isp    = "—"
    wans   = rep.get("wans", [])
    if isinstance(wans, list) and wans and isinstance(wans[0], dict):
        isp = wans[0].get("ispInfo", {}).get("name", "—")
    uptime  = "—"
    periods = rep.get("internetIssues5min", {}).get("periods", [])
    if isinstance(periods, list) and periods:
        last = periods[-1]
        if isinstance(last, dict) and last.get("wanUptime") is not None:
            uptime = f"{last['wanUptime']:.1f}%"
    return {"nome": nome, "estado": estado, "ip": ip, "isp": isp, "uptime": uptime, "ultimo_sinal": ultimo_sinal}

def buscar_ineps_ubiquiti(api_key: str, apelido: str, ineps: set) -> dict:
    out   = {}
    hosts = get_paginated_hosts(api_key)
    for h in hosts:
        d    = extrair_host(h)
        nome = d["nome"].upper()
        for inep in ineps:
            if str(inep).strip().upper() in nome:
                status_raw = "ONLINE" if d["estado"] == "connected" else "OFFLINE"
                inep_key = str(inep).strip()
                out.setdefault(inep_key, []).append({
                    "Status Rede":     f"UBIQUITI - {status_raw}",
                    "Plataforma":      "UBIQUITI",
                    LAST_SEEN_COLUMN:   d["ultimo_sinal"],
                    "Uptime WAN":      d["uptime"],
                    "ISP":             d["isp"],
                    "Conta":           apelido,
                    "IP Externo":      d["ip"],
                    "Nome no Console": d["nome"],
                })
                break
    return out

def coletar_todos_hosts_ubiquiti(contas: list) -> tuple:
    """Retorna TODOS os hosts de todas as contas, sem filtro por INEP."""
    rows  = []
    erros = []
    for c in contas:
        if not c.get("api_key", "").strip():
            continue
        try:
            hosts = get_paginated_hosts(c["api_key"])
            for h in hosts:
                d      = extrair_host(h)
                inep   = extrair_inep_do_nome(d["nome"])
                status = "ONLINE" if d["estado"] == "connected" else "OFFLINE"
                rows.append({
                    "INEP":            inep or "—",
                    "Nome no Console": d["nome"],
                    "Status Rede":     f"UBIQUITI - {status}",
                    "Plataforma":      "UBIQUITI",
                    LAST_SEEN_COLUMN:   d["ultimo_sinal"],
                    "Uptime WAN":      d["uptime"],
                    "ISP":             d["isp"],
                    "Conta":           c["apelido"],
                    "IP Externo":      d["ip"],
                })
        except Exception as e:
            erros.append(f"{c['apelido']}: {e}")
    return rows, erros

def coletar_todos_isps_ubiquiti(contas: list) -> tuple:
    isps  = set()
    erros = []
    for c in contas:
        if not c.get("api_key", "").strip():
            continue
        try:
            hosts = get_paginated_hosts(c["api_key"])
            for h in hosts:
                d = extrair_host(h)
                if d["isp"] and d["isp"] not in ("—", ""):
                    isps.add(d["isp"].strip())
        except Exception as e:
            erros.append(f"{c['apelido']}: {e}")
    return isps, erros

def testar_conta(api_key: str):
    try:
        r = requests.get(
            f"{BASE_URL}/hosts",
            headers={"X-API-KEY": api_key}, params={"pageSize": 1}, timeout=10
        )
        if r.status_code == 200:
            return True, f"OK — {len(r.json().get('data', []))} host(s) retornado(s)"
        return False, f"HTTP {r.status_code}"
    except requests.exceptions.Timeout:
        return False, "Timeout"
    except Exception as e:
        return False, str(e)

# ═══════════════════════════════════════════════════════════════
# OMADA — PROCESSAMENTO DO EXPORT
# ═══════════════════════════════════════════════════════════════
def extrair_inep_do_nome(nome: str):
    matches = re.findall(r'\b(\d{8})\b', str(nome))
    return matches[-1] if matches else None

def processar_export_omada(df_omada: pd.DataFrame) -> dict:
    resultado = {}
    df_omada.columns = [str(c).strip().upper() for c in df_omada.columns]
    if "NAME" not in df_omada.columns or "STATUS" not in df_omada.columns:
        return resultado
    for _, row in df_omada.iterrows():
        nome   = str(row.get("NAME", "")).strip()
        status = str(row.get("STATUS", "")).strip().upper()
        inep   = extrair_inep_do_nome(nome)
        if not inep:
            continue
        status_raw = "ONLINE" if status == "ONLINE" else "OFFLINE"
        ip_externo = "—"
        if "IP ADDRESS" in df_omada.columns:
            ips = str(row.get("IP ADDRESS", "")).split(",")
            ip_externo = ips[-1].strip() if ips else "—"
        resultado[inep] = {
            "Status Rede":     f"OMADA - {status_raw}",
            "Plataforma":      "OMADA",
            LAST_SEEN_COLUMN:   "—",
            "Uptime WAN":      "—",
            "ISP":             "—",
            "Conta":           "Omada Cloud",
            "IP Externo":      ip_externo,
            "Nome no Console": nome,
        }
    return resultado

def processar_export_omada_completo(df_omada: pd.DataFrame) -> list:
    """Versão para inventário — retorna lista de todas as linhas, não só cruzadas."""
    rows = []
    df_omada.columns = [str(c).strip().upper() for c in df_omada.columns]
    if "NAME" not in df_omada.columns or "STATUS" not in df_omada.columns:
        return rows
    for _, row in df_omada.iterrows():
        nome   = str(row.get("NAME", "")).strip()
        status = str(row.get("STATUS", "")).strip().upper()
        inep   = extrair_inep_do_nome(nome)
        status_raw = "ONLINE" if status == "ONLINE" else "OFFLINE"
        ip_externo = "—"
        if "IP ADDRESS" in df_omada.columns:
            ips = str(row.get("IP ADDRESS", "")).split(",")
            ip_externo = ips[-1].strip() if ips else "—"
        rows.append({
            "INEP":            inep or "—",
            "Nome no Console": nome,
            "Status Rede":     f"OMADA - {status_raw}",
            "Plataforma":      "OMADA",
            LAST_SEEN_COLUMN:   "—",
            "Uptime WAN":      "—",
            "ISP":             "—",
            "Conta":           "Omada Cloud",
            "IP Externo":      ip_externo,
        })
    return rows

# ═══════════════════════════════════════════════════════════════
# ZYXEL — PROCESSAMENTO DO EXPORT CSV
# ═══════════════════════════════════════════════════════════════
ZYXEL_STATUS_MAP = {
    "OK":                  "ONLINE",
    "DEVICE ALERTED":      "ALERTA",
    "DEVICE OFFLINE":      "OFFLINE",
    "DEVICES UNREACHABLE": "OFFLINE",
    "NO DEVICES":          "SEM DISPOSITIVO",
}

def processar_export_zyxel(df_zyxel: pd.DataFrame) -> dict:
    resultado = {}
    col_map   = {}
    for col in df_zyxel.columns:
        cu = str(col).strip().upper()
        if cu in ("ESTADO", "STATUS"):               col_map["status"]      = col
        elif cu in ("NOME", "NAME"):                  col_map["nome"]        = col
        elif cu in ("DISPOSITIVOS OFFLINE", "OFFLINE DEVICES"): col_map["offline"] = col
        elif cu in ("DISPOSITIVOS", "DEVICES"):       col_map["total_dev"]   = col
        elif cu in ("% OFFLINE", "% OFFLINE DEVICES"):col_map["pct_offline"] = col
    if "status" not in col_map or "nome" not in col_map:
        return resultado
    for _, row in df_zyxel.iterrows():
        nome           = str(row[col_map["nome"]]).strip()
        status_raw     = str(row[col_map["status"]]).strip().upper()
        inep           = extrair_inep_do_nome(nome)
        if not inep: continue
        status_interno = ZYXEL_STATUS_MAP.get(status_raw, "OFFLINE")
        total_dev      = str(row.get(col_map.get("total_dev",  ""), "—")).strip()
        offline_dev    = str(row.get(col_map.get("offline",    ""), "—")).strip()
        pct_offline    = str(row.get(col_map.get("pct_offline",""), "—")).strip()
        resultado[inep] = {
            "Status Rede":       f"ZYXEL - {status_interno}",
            "Plataforma":        "ZYXEL",
            LAST_SEEN_COLUMN:     "—",
            "Uptime WAN":        f"{pct_offline} offline" if pct_offline != "—" else "—",
            "ISP":               "—",
            "Conta":             "Zyxel Nebula",
            "IP Externo":        "—",
            "Nome no Console":   nome,
            "Devices (tot/off)": f"{total_dev} / {offline_dev}",
        }
    return resultado

def processar_export_zyxel_completo(df_zyxel: pd.DataFrame) -> list:
    """Versão para inventário — retorna todas as linhas."""
    rows    = []
    col_map = {}
    for col in df_zyxel.columns:
        cu = str(col).strip().upper()
        if cu in ("ESTADO", "STATUS"):                col_map["status"]      = col
        elif cu in ("NOME", "NAME"):                   col_map["nome"]        = col
        elif cu in ("DISPOSITIVOS OFFLINE", "OFFLINE DEVICES"): col_map["offline"] = col
        elif cu in ("DISPOSITIVOS", "DEVICES"):        col_map["total_dev"]   = col
        elif cu in ("% OFFLINE", "% OFFLINE DEVICES"): col_map["pct_offline"] = col
    if "status" not in col_map or "nome" not in col_map:
        return rows
    for _, row in df_zyxel.iterrows():
        nome           = str(row[col_map["nome"]]).strip()
        status_raw     = str(row[col_map["status"]]).strip().upper()
        inep           = extrair_inep_do_nome(nome)
        status_interno = ZYXEL_STATUS_MAP.get(status_raw, "OFFLINE")
        total_dev      = str(row.get(col_map.get("total_dev",  ""), "—")).strip()
        offline_dev    = str(row.get(col_map.get("offline",    ""), "—")).strip()
        pct_offline    = str(row.get(col_map.get("pct_offline",""), "—")).strip()
        rows.append({
            "INEP":              inep or "—",
            "Nome no Console":   nome,
            "Status Rede":       f"ZYXEL - {status_interno}",
            "Plataforma":        "ZYXEL",
            LAST_SEEN_COLUMN:     "—",
            "Uptime WAN":        f"{pct_offline} offline" if pct_offline != "—" else "—",
            "ISP":               "—",
            "Conta":             "Zyxel Nebula",
            "IP Externo":        "—",
            "Devices (tot/off)": f"{total_dev} / {offline_dev}",
        })
    return rows

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════
COR_STATUS = {
    "UBIQUITI - ONLINE":       "background:#052e16; color:#22c55e; font-weight:600",
    "UBIQUITI - OFFLINE":      "background:#2d0a0a; color:#ef4444; font-weight:600",
    "OMADA - ONLINE":          "background:#0a1f2e; color:#38bdf8; font-weight:600",
    "OMADA - OFFLINE":         "background:#2d1a0a; color:#fb923c; font-weight:600",
    "ZYXEL - ONLINE":          "background:#1a0a2e; color:#a855f7; font-weight:600",
    "ZYXEL - OFFLINE":         "background:#2d0a1f; color:#f472b6; font-weight:600",
    "ZYXEL - ALERTA":          "background:#2d2200; color:#facc15; font-weight:600",
    "ZYXEL - SEM DISPOSITIVO": "background:#1a1f2e; color:#6b7280; font-weight:600",
    "NÃO ENCONTRADO":          "background:#1a1f2e; color:#374151; font-weight:600",
}

def cor_status(val):
    return COR_STATUS.get(str(val).strip(), "")

def cor_estado(val):
    if val == "connected":    return "color:#22c55e; font-weight:600"
    if val == "disconnected": return "color:#ef4444; font-weight:600"
    return ""

def to_xlsx(df: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Resultado")
    return buf.getvalue()


def _extrair_uf(nome: str) -> str:
    m = re.match(r"^([A-Z]{2})\s", str(nome).strip().upper())
    return m.group(1) if m else "??"


@st.cache_data(show_spinner=False)
def gerar_excel_inventario_formatado(df_inv: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Paleta de cores ──────────────────────────────────────────
    NAVY    = "1F4E79"
    BLUE    = "2E75B6"
    PURPLE  = "7030A0"
    RED_HDR = "C00000"
    WHITE   = "FFFFFF"
    GRN_BG  = "E2EFDA"; GRN_FG = "375623"
    RED_BG  = "FCE4D6"; RED_FG = "C00000"
    ORG_BG  = "FFF2CC"; ORG_FG = "7F6000"
    BLU_BG  = "DEEAF1"
    GRAY    = "F2F2F2"
    BD      = "BFBFBF"

    def _side():   return Side(style="thin", color=BD)
    def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())

    def _cell(ws, row, col, value=None, bg=None, fg="000000", bold=False,
              size=10, align="left", italic=False, border=True):
        c = ws.cell(row=row, column=col)
        if value is not None:
            c.value = value
        c.font = Font(name="Calibri", size=size, color=fg, bold=bold, italic=italic)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if border:
            c.border = _border()
        return c

    def _hdr(ws, row, col, value, bg=NAVY):
        return _cell(ws, row, col, value, bg=bg, fg=WHITE, bold=True,
                     size=11, align="center")

    def _title(ws, row, n_cols, text, bg=NAVY, size=14, height=38):
        _cell(ws, row, 1, text, bg=bg, fg=WHITE, bold=True,
              size=size, align="center", border=False)
        if n_cols > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=n_cols)
        ws.row_dimensions[row].height = height

    def _merge_row(ws, row, col_start, col_end):
        if col_end > col_start:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_end)

    def _auto_width(ws, min_w=10, max_w=55):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[letter].width = min(max(width + 3, min_w), max_w)

    # ── Preparar dados ───────────────────────────────────────────
    df = df_inv.copy()
    if "Status Rede" in df.columns:
        df["Status"] = df["Status Rede"].apply(
            lambda x: "Online" if "ONLINE" in str(x).upper() else "Offline"
        )
    elif "Status" not in df.columns:
        df["Status"] = "Offline"

    df["INEP"] = df["INEP"].astype(str).str.strip()
    _f42 = {"Omada Cloud", "Zyxel Nebula"}
    df["FASE"] = df["Conta"].apply(lambda c: "4.2" if c in _f42 else "") \
        if "Conta" in df.columns else ""

    col_base = [c for c in ["INEP", "Nome no Console", "Status", "Conta", "FASE"]
                if c in df.columns]
    df = df[col_base].fillna("").copy()

    # Referência
    ineps_ref: set = set()
    if os.path.exists("data/lista_ineps_referencia.xlsx"):
        try:
            _dr = pd.read_excel("data/lista_ineps_referencia.xlsx", dtype=str)
            _ci = next((c for c in _dr.columns if "INEP" in c.upper()), _dr.columns[0])
            ineps_ref = set(_dr[_ci].dropna().astype(str).str.strip())
        except Exception:
            pass

    ineps_enc = set(df[df["INEP"] != "—"]["INEP"].unique())
    nao_enc   = sorted(ineps_ref - ineps_enc)
    df_nao    = pd.DataFrame({"Ordem": range(1, len(nao_enc)+1), "INEP": nao_enc})

    validos = df[df["INEP"] != "—"]
    total   = len(validos)
    online  = int((validos["Status"] == "Online").sum())  if "Status" in validos.columns else 0
    offline = int((validos["Status"] == "Offline").sum()) if "Status" in validos.columns else 0

    if not validos.empty and "Nome no Console" in validos.columns:
        _uf_s = validos["Nome no Console"].apply(_extrair_uf)
        por_uf = (validos.assign(_uf=_uf_s)
                  .groupby("_uf")
                  .agg(Total=("INEP","count"),
                       Onl=("Status", lambda x:(x=="Online").sum()),
                       Off=("Status", lambda x:(x=="Offline").sum()))
                  .reset_index()
                  .rename(columns={"_uf":"UF","Onl":"Online","Off":"Offline"})
                  .sort_values("Total", ascending=False))
    else:
        por_uf = pd.DataFrame(columns=["UF","Total","Online","Offline"])

    contas_ordem = sorted(df["Conta"].dropna().unique().tolist()) \
        if "Conta" in df.columns else []
    df_42 = df[df["FASE"] == "4.2"].copy() if "FASE" in df.columns \
        else pd.DataFrame(columns=col_base)

    # ── Workbook ─────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ════════════════════════════════════════════════════════════
    # ABA: ESTATISTICAS
    # ════════════════════════════════════════════════════════════
    ws = wb.create_sheet("ESTATISTICAS")
    ws.sheet_view.showGridLines = False
    NC = 5  # colunas A–E

    # Linha 1 — título principal
    _title(ws, 1, NC, "INVENTÁRIO GERAL DE ESCOLAS — EACE", bg=NAVY, size=16, height=44)

    # Linha 2 — data
    _cell(ws, 2, 1,
          f"Gerado em {datetime.now(APP_TZ).strftime('%d/%m/%Y às %H:%M')} (Brasília)",
          fg="595959", italic=True, align="center", border=False)
    _merge_row(ws, 2, 1, NC)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10  # espaçador

    # Linha 4 — cabeçalho da seção KPI
    _title(ws, 4, NC, "INDICADORES GERAIS", bg=BLUE, size=12, height=26)

    # Linha 5 — colunas da tabela KPI
    _hdr(ws, 5, 1, "INDICADOR", bg=NAVY)
    _hdr(ws, 5, 2, "VALOR",     bg=NAVY)
    _merge_row(ws, 5, 2, NC)
    ws.row_dimensions[5].height = 22

    # Linhas 6–10 — KPIs
    kpis = [
        ("Total de escolas no inventário",        total,   BLU_BG, "000000"),
        ("Online",                                 online,  GRN_BG, GRN_FG),
        ("Offline",                                offline, RED_BG, RED_FG),
        ("Percentual online",
         f"{online/total*100:.2f}%" if total else "—",
         GRN_BG if (total and online/total >= .5) else RED_BG, "000000"),
        ("Não encontradas em nenhum dashboard",   len(nao_enc), ORG_BG, ORG_FG),
    ]
    for i, (label, val, bg, fg) in enumerate(kpis):
        r = 6 + i
        _cell(ws, r, 1, label, bg=bg, fg=fg, size=11)
        _cell(ws, r, 2, val,   bg=bg, fg=fg, size=12, bold=True, align="center")
        _merge_row(ws, r, 2, NC)
        ws.row_dimensions[r].height = 24

    ws.row_dimensions[11].height = 12  # espaçador

    # Linha 12 — cabeçalho UF
    _title(ws, 12, NC, "RESUMO POR ESTADO (UF)", bg=NAVY, size=12, height=28)

    # Linha 13 — colunas UF
    for ci, h in enumerate(["UF", "Total", "Online", "Offline", "% Online"], 1):
        _hdr(ws, 13, ci, h, bg=BLUE)
    ws.row_dimensions[13].height = 22

    # Linhas 14+ — dados UF
    for i, (_, r) in enumerate(por_uf.iterrows()):
        ri = 14 + i
        pct = f"{r['Online']/r['Total']*100:.1f}%" if r["Total"] else "—"
        bg_row = GRAY if i % 2 == 0 else WHITE
        for ci, v in enumerate([r["UF"], int(r["Total"]), int(r["Online"]),
                                 int(r["Offline"]), pct], 1):
            _cell(ws, ri, ci, v, bg=bg_row, align="center")
        ws.row_dimensions[ri].height = 18

    ws.column_dimensions["A"].width = 40
    for letter in ["B","C","D","E"]:
        ws.column_dimensions[letter].width = 14

    # ════════════════════════════════════════════════════════════
    # Helper — criar aba de dados com formatação completa
    # ════════════════════════════════════════════════════════════
    def criar_aba(nome, df_data, titulo, bg_titulo=NAVY):
        ws2 = wb.create_sheet(nome)
        ws2.sheet_view.showGridLines = False

        cols = list(df_data.columns)
        nc   = len(cols) or 1

        # Linha 1 — título
        _title(ws2, 1, nc, titulo, bg=bg_titulo, size=13, height=36)

        if df_data.empty:
            _cell(ws2, 2, 1, "Sem dados disponíveis.", italic=True,
                  fg="595959", border=False)
            ws2.column_dimensions["A"].width = 40
            return

        # Linha 2 — cabeçalhos
        CENTRAR = {"INEP", "Status", "Conta", "FASE", "Ordem"}
        for ci, col_name in enumerate(cols, 1):
            _hdr(ws2, 2, ci, col_name, bg=NAVY)
        ws2.row_dimensions[2].height = 26

        # Linhas 3+ — dados
        for i, (_, row_data) in enumerate(df_data.iterrows()):
            ri     = i + 3
            status = str(row_data.get("Status", "")).strip()

            if status == "Online":
                bg_row, fg_row = GRN_BG, GRN_FG
            elif status == "Offline":
                bg_row, fg_row = RED_BG, RED_FG
            else:
                bg_row = GRAY if i % 2 == 0 else WHITE
                fg_row = "000000"

            for ci, col_name in enumerate(cols, 1):
                val = row_data.get(col_name, "")
                if pd.isnull(val):
                    val = ""
                _cell(ws2, ri, ci, val, bg=bg_row, fg=fg_row,
                      align="center" if col_name in CENTRAR else "left")

            ws2.row_dimensions[ri].height = 18

        # Freeze após título + cabeçalho; filtro automático
        ws2.freeze_panes = "A3"
        ws2.auto_filter.ref = f"A2:{get_column_letter(nc)}2"
        _auto_width(ws2)

    # ── Criar todas as abas ──────────────────────────────────────
    criar_aba("Geral", df,
              "INVENTÁRIO GERAL — TODAS AS PLATAFORMAS")

    criar_aba("Escolas não encontradas dash´s", df_nao,
              f"ESCOLAS NÃO ENCONTRADAS EM NENHUM DASHBOARD  ({len(nao_enc)} INEPs)",
              bg_titulo=RED_HDR)

    for conta in contas_ordem:
        df_c  = df[df["Conta"] == conta].copy()
        sname = re.sub(r'[\\/*?:\[\]]', "_", conta)[:31]
        bg    = PURPLE if conta in _f42 else NAVY
        criar_aba(sname, df_c, f"ESCOLAS — {conta.upper()}", bg_titulo=bg)

    criar_aba("4.2", df_42,
              "FASE 4.2 — OMADA + ZYXEL", bg_titulo=PURPLE)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def deduplicar_por_inep(todos: dict) -> dict:
    """
    Recebe dict {inep: dados} ou {inep: [dados]} e garante que cada INEP
    fique com a entrada de maior qualidade.
    Prioridade: ONLINE > tem ISP > tem IP > tem Uptime.
    INEPs '—' são ignorados (não há como deduplicar sem chave).
    """
    def _score(dados):
        s = 0
        if "ONLINE" in str(dados.get("Status Rede", "")): s += 100
        if str(dados.get("ISP",        "—")) not in ("—", "", "nan"): s += 10
        if str(dados.get("IP Externo", "—")) not in ("—", "", "nan"): s += 5
        if str(dados.get("Uptime WAN", "—")) not in ("—", "", "nan"): s += 1
        return s

    # Agrupa candidatos por INEP
    candidatos: dict[str, list] = {}
    for inep, dados in todos.items():
        if isinstance(dados, list):
            candidatos.setdefault(inep, []).extend(dados)
        else:
            candidatos.setdefault(inep, []).append(dados)

    resultado = {}
    for inep, lista in candidatos.items():
        melhor = max(lista, key=_score)
        resultado[inep] = melhor
    return resultado


def adicionar_candidatos_por_inep(destino: dict, origem: dict, ineps_filtrados: set | None = None) -> None:
    """Acumula todas as ocorrências para permitir escolher ONLINE mesmo vindo de outra conta."""
    for inep, dados in origem.items():
        inep_key = str(inep).strip()
        if ineps_filtrados is not None and inep_key not in ineps_filtrados:
            continue
        itens = dados if isinstance(dados, list) else [dados]
        destino.setdefault(inep_key, []).extend(itens)


def deduplicar_df_por_inep(df: pd.DataFrame) -> tuple:
    """
    Versão DataFrame para o Inventário Geral.
    Retorna (df_deduplicado, n_removidos).
    """
    def _score_row(row):
        s = 0
        if "ONLINE" in str(row.get("Status Rede", "")): s += 100
        if str(row.get("ISP",        "—")) not in ("—", "", "nan"): s += 10
        if str(row.get("IP Externo", "—")) not in ("—", "", "nan"): s += 5
        if str(row.get("Uptime WAN", "—")) not in ("—", "", "nan"): s += 1
        return s

    sem_inep = df[df["INEP"] == "—"].copy()
    com_inep = df[df["INEP"] != "—"].copy()
    if not com_inep.empty:
        com_inep["_score"] = com_inep.apply(_score_row, axis=1)
        com_inep = (
            com_inep
            .sort_values("_score", ascending=False)
            .drop_duplicates(subset=["INEP"], keep="first")
            .drop(columns=["_score"])
        )
    antes     = len(df)
    df_result = pd.concat([com_inep, sem_inep], ignore_index=True)
    return df_result, antes - len(df_result)


def mesclar_isps_novos(isps_novos: set) -> int:
    provs = dict(st.session_state.provedores)
    novos = 0
    for isp in isps_novos:
        isp = str(isp).strip()
        if isp and isp not in ("—", "nan", "") and isp not in provs:
            provs[isp] = {"telefone": "", "celular": "", "observacao": ""}
            novos += 1
    if novos > 0:
        st.session_state.provedores = provs
        salvar_provedores(provs)
    return novos

# ═══════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    st.caption(f"Conectado: `{st.session_state.hub_user_email}`")
    if st.button("Sair", use_container_width=True):
        st.session_state.hub_authenticated = False
        st.session_state.hub_user_email = ""
        st.session_state.daily_report_authenticated = False
        st.rerun()

    st.markdown('<hr class="saas-divider">', unsafe_allow_html=True)

    # Contas carregadas dos Secrets — sem exposição de chaves na UI
    validas = [c for c in st.session_state.contas if c.get("api_key", "").strip()]

    st.markdown('<div class="sidebar-label">Contas Ubiquiti</div>', unsafe_allow_html=True)
    if validas:
        sel  = st.multiselect(
            "Selecionar contas:",
            [c["apelido"] for c in validas],
            default=[c["apelido"] for c in validas],
            label_visibility="collapsed"
        )
        alvo = [c for c in validas if c["apelido"] in sel]
    else:
        st.warning("Nenhuma conta configurada nos Secrets.")
        alvo = []

        # Diagnóstico — ajuda a identificar problema de formato
        with st.expander("🔍 Diagnóstico Secrets", expanded=True):
            try:
                todas = dict(st.secrets)
                chaves = list(todas.keys())
                st.caption(f"Chaves encontradas: `{chaves}`")
                if "contas" in st.secrets:
                    st.caption(f"✅ Seção [contas] encontrada com {len(dict(st.secrets['contas']))} entradas.")
                else:
                    st.caption("⚠️ Seção [contas] não encontrada — tentando ler chaves soltas na raiz.")
                    candidatas = [k for k in chaves if any(k.startswith(p) for p in _CONTA_PREFIXOS)
                                  or (len(str(todas[k])) >= 20 and " " not in str(todas[k]))]
                    st.caption(f"Chaves candidatas a conta: `{candidatas}`")
                    if candidatas:
                        st.caption(f"✅ {len(candidatas)} conta(s) detectada(s) pelo formato alternativo.")
                    else:
                        st.caption("❌ Nenhuma conta detectada. Use o formato abaixo:")
                        st.code("[contas]\nConta_AC1 = \"sua_chave\"", language="toml")
            except Exception as ex:
                st.caption(f"Erro ao ler Secrets: {ex}")

    render_daily_closing_admin()

    st.markdown('<div class="sidebar-label">Status</div>', unsafe_allow_html=True)
    ca, cb = st.columns(2)
    ca.metric("Contas",  len(validas))
    cb.metric("Ativas",  len(alvo))
    if st.session_state.ultima_consulta:
        st.caption(f"Última consulta: {st.session_state.ultima_consulta}")

# ═══════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════
st.markdown(
    '<p class="page-title">Hub Redes — EACE</p>'
    '<p class="page-sub">Monitor de Conectividade · Ubiquiti · Omada · Zyxel · Escolas Brasileiras</p>',
    unsafe_allow_html=True
)
st.markdown('<hr class="saas-divider">', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# ABAS
# ═══════════════════════════════════════════════════════════════
t1, t2, t3, t4, t5, t6 = st.tabs([
    "📄 Relatório Diário",
    "🌐 Inventário Geral",
    "🔍 Busca Manual",
    "🛠️ Dados das contas",
    "🩺 Diagnóstico das API´S",
    "ℹ️ Ajuda",
])

# ─────────────────────────────────────────
# ABA 1 - RELATORIO DIARIO
# ─────────────────────────────────────────
with t1:
    render_daily_report(st.session_state.daily_report_current)

# ─────────────────────────────────────────
# ABA 2 - INVENTARIO GERAL
# ─────────────────────────────────────────

# ─────────────────────────────────────────
with t2:
    st.markdown("### 🌐 Inventário Geral de Escolas")
    st.write(
        "Coleta **todas** as escolas de todas as plataformas, independente de chamados abertos. "
        "Ubiquiti é buscado via API. Omada e Zyxel via upload dos exports."
    )

    st.markdown('<hr class="saas-divider">', unsafe_allow_html=True)

    # Uploads Omada e Zyxel para o inventário
    col_inv_om, col_inv_zy = st.columns(2)
    with col_inv_om:
        st.caption("🔵 **Omada** — export do portal On Premise Systems")
        inv_omada = st.file_uploader("Export Omada (.xlsx)", type=["xlsx"], key="inv_omada")
        if inv_omada:
            df_inv_om_prev = pd.read_excel(inv_omada)
            inv_omada.seek(0)
            st.caption(f"{len(df_inv_om_prev)} controllers carregados")
    with col_inv_zy:
        st.caption("🟣 **Zyxel** — export CSV do Nebula (Overview → Sites)")
        inv_zyxel = st.file_uploader("Export Zyxel (.csv)", type=["csv"], key="inv_zyxel")
        if inv_zyxel:
            df_inv_zy_prev = pd.read_csv(inv_zyxel)
            inv_zyxel.seek(0)
            st.caption(f"{len(df_inv_zy_prev)} sites carregados")

    st.markdown('<hr class="saas-divider">', unsafe_allow_html=True)

    col_btn_inv, col_reset_inv, _ = st.columns([2, 2, 4])
    with col_btn_inv:
        if st.button("🔄 Coletar Inventário Completo", type="primary", use_container_width=True):
            if not alvo and not inv_omada and not inv_zyxel:
                st.warning("Configure ao menos uma fonte: conta Ubiquiti, export Omada ou export Zyxel.")
            else:
                todas_linhas = []
                log_inv      = []
                t0           = time.time()

                # ── Ubiquiti — todas as contas em paralelo ──
                if alvo:
                    prog_inv = st.progress(0.0, text="Coletando Ubiquiti...")
                    with ThreadPoolExecutor(max_workers=min(len(alvo), 5)) as ex:
                        futs_inv = {
                            ex.submit(coletar_todos_hosts_ubiquiti, [c]): c["apelido"]
                            for c in alvo
                        }
                        done_inv = 0
                        for fut in as_completed(futs_inv):
                            ap = futs_inv[fut]
                            try:
                                rows_c, erros_c = fut.result()
                                todas_linhas.extend(rows_c)
                                for e in erros_c: log_inv.append(("err", e))
                                log_inv.append(("ok", f"Ubiquiti {ap}: {len(rows_c)} host(s)"))
                            except Exception as e:
                                log_inv.append(("err", f"Ubiquiti {ap}: {e}"))
                            done_inv += 1
                            prog_inv.progress(done_inv / len(alvo), text=f"Ubiquiti {done_inv}/{len(alvo)}...")
                    prog_inv.progress(1.0, text="Ubiquiti concluído.")

                # ── Omada ──
                if inv_omada:
                    try:
                        inv_omada.seek(0)
                        rows_om = processar_export_omada_completo(pd.read_excel(inv_omada))
                        todas_linhas.extend(rows_om)
                        log_inv.append(("ok", f"Omada: {len(rows_om)} controller(s)"))
                    except Exception as e:
                        log_inv.append(("err", f"Omada: {e}"))

                # ── Zyxel ──
                if inv_zyxel:
                    try:
                        inv_zyxel.seek(0)
                        rows_zy = processar_export_zyxel_completo(pd.read_csv(inv_zyxel))
                        todas_linhas.extend(rows_zy)
                        log_inv.append(("ok", f"Zyxel: {len(rows_zy)} site(s)"))
                    except Exception as e:
                        log_inv.append(("err", f"Zyxel: {e}"))

                if todas_linhas:
                    df_inv = pd.DataFrame(todas_linhas)

                    # Deduplicação centralizada
                    df_inv, removidos = deduplicar_df_por_inep(df_inv)
                    if removidos > 0:
                        log_inv.append(("ok", f"Deduplicação: {removidos} duplicata(s) removida(s)"))

                    # Garante ordem das colunas
                    col_order = ["INEP","Nome no Console","Status Rede","Plataforma",
                                 LAST_SEEN_COLUMN,"Uptime WAN","ISP","Conta","IP Externo"]
                    df_inv = df_inv[[c for c in col_order if c in df_inv.columns] +
                                    [c for c in df_inv.columns if c not in col_order]]
                    st.session_state.df_inventario     = df_inv
                    st.session_state.inv_ultima_coleta = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                    st.success(f"Inventário coletado em {time.time()-t0:.1f}s — {len(df_inv):,} escola(s) únicas.")

                    # Log
                    html_log = "".join(
                        f'<div class="log-{"ok" if tp=="ok" else "err"}">{"✓" if tp=="ok" else "✗"}  {m}</div>'
                        for tp, m in log_inv
                    )
                    st.markdown(f'<div class="log-box">{html_log}</div>', unsafe_allow_html=True)
                else:
                    st.error("Nenhuma escola coletada. Verifique as fontes.")

    with col_reset_inv:
        if st.button("🗑️ Limpar Inventário", use_container_width=True):
            st.session_state.df_inventario     = None
            st.session_state.inv_ultima_coleta = None
            st.rerun()

    # ── Exibição do inventário ──
    if st.session_state.df_inventario is not None:
        df_inv = st.session_state.df_inventario
        st.markdown(f'<div class="ts-pill">🕐 {st.session_state.inv_ultima_coleta}</div>', unsafe_allow_html=True)

        # Cards
        total_inv   = len(df_inv)
        online_inv  = df_inv["Status Rede"].str.contains("ONLINE",   na=False).sum()
        offline_inv = df_inv["Status Rede"].str.contains("OFFLINE",  na=False).sum()
        n_ubi_inv   = df_inv["Status Rede"].str.contains("UBIQUITI", na=False).sum()
        n_om_inv    = df_inv["Status Rede"].str.contains("OMADA",    na=False).sum()
        n_zy_inv    = df_inv["Status Rede"].str.contains("ZYXEL",    na=False).sum()

        st.markdown(f"""
        <div class="saas-grid">
          <div class="saas-card"><div class="saas-card-label">Total escolas</div>
            <div class="saas-card-value c-blue">{total_inv:,}</div><div class="saas-card-sub">todas as plataformas</div></div>
          <div class="saas-card"><div class="saas-card-label">Online</div>
            <div class="saas-card-value c-green">{online_inv:,}</div>
            <div class="saas-card-sub">{f"{online_inv/total_inv*100:.0f}%" if total_inv else "—"}</div></div>
          <div class="saas-card"><div class="saas-card-label">Offline / Alerta</div>
            <div class="saas-card-value c-red">{offline_inv:,}</div>
            <div class="saas-card-sub">{f"{offline_inv/total_inv*100:.0f}%" if total_inv else "—"}</div></div>
          <div class="saas-card"><div class="saas-card-label">Ubiquiti</div>
            <div class="saas-card-value c-blue">{n_ubi_inv:,}</div><div class="saas-card-sub">hosts</div></div>
          <div class="saas-card"><div class="saas-card-label">Omada</div>
            <div class="saas-card-value c-teal">{n_om_inv:,}</div><div class="saas-card-sub">controllers</div></div>
          <div class="saas-card"><div class="saas-card-label">Zyxel</div>
            <div class="saas-card-value c-purple">{n_zy_inv:,}</div><div class="saas-card-sub">sites</div></div>
        </div>""", unsafe_allow_html=True)

        # Pesquisa e filtros
        with st.expander("⚙️ Filtros e Pesquisa", expanded=True):
            pesq_col, f1_inv, f2_inv, f3_inv = st.columns([2, 1, 1, 1])

            pesq = pesq_col.text_input(
                "🔎 Pesquisar INEP ou nome da escola:",
                placeholder="Ex: 23000066 ou ACARAÚ",
                key="inv_pesq"
            )
            op_plat_inv = sorted(df_inv["Plataforma"].dropna().unique().tolist())
            op_st_inv   = sorted(df_inv["Status Rede"].dropna().unique().tolist())
            op_isp_inv  = sorted([v for v in df_inv["ISP"].dropna().unique().tolist() if v not in ("—","")])

            fp_inv  = f1_inv.multiselect("Plataforma", op_plat_inv, default=[], key="fp_inv")
            fs_inv  = f2_inv.multiselect("Status",     op_st_inv,   default=[], key="fs_inv")
            fi_inv  = f3_inv.multiselect("ISP",        op_isp_inv,  default=[], key="fi_inv")

        df_inv_f = df_inv.copy()

        # Pesquisa textual em INEP e Nome
        if pesq.strip():
            mask = (
                df_inv_f["INEP"].astype(str).str.contains(pesq.strip(), case=False, na=False) |
                df_inv_f["Nome no Console"].astype(str).str.contains(pesq.strip(), case=False, na=False)
            )
            df_inv_f = df_inv_f[mask]

        if fp_inv: df_inv_f = df_inv_f[df_inv_f["Plataforma"].isin(fp_inv)]
        if fs_inv: df_inv_f = df_inv_f[df_inv_f["Status Rede"].isin(fs_inv)]
        if fi_inv: df_inv_f = df_inv_f[df_inv_f["ISP"].isin(fi_inv)]

        st.dataframe(
            df_inv_f.style.map(cor_status, subset=["Status Rede"]),
            use_container_width=True,
            height=520
        )
        st.caption(f"{len(df_inv_f):,} de {len(df_inv):,} escolas exibidas")

        ce1, ce2, _ = st.columns([1, 1, 3])
        ce1.download_button(
            "⬇️ Exportar inventário completo",
            gerar_excel_inventario_formatado(df_inv),
            f"inventario_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Planilha formatada: ESTATISTICAS, Geral, Escolas não encontradas, por conta."
        )
        ce2.download_button(
            "⬇️ Exportar visualização atual",
            to_xlsx(df_inv_f),
            f"inventario_filtrado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Exporta somente o que está visível na tabela (com filtros aplicados)."
        )
    else:
        st.info("Clique em **Coletar Inventário Completo** para carregar todas as escolas.")


# ─────────────────────────────────────────
# ABA 3 — BUSCA MANUAL
# ─────────────────────────────────────────
with t3:
    st.markdown("#### INEPs para consulta")
    col_txt, col_how = st.columns([2, 1])
    with col_txt:
        txt = st.text_area(
            "INEPs (um por linha ou separados por vírgula):",
            height=160, placeholder="13084259\n13051997\n15559556"
        )
    with col_how:
        st.markdown("**Como usar**")
        st.markdown(
            "- Um INEP por linha ou separados por vírgula\n"
            "- Busca em todas as fontes disponíveis\n"
            "- Ubiquiti: contas ativas na barra lateral\n"
            "- Omada/Zyxel: faça upload dos exports abaixo\n"
            "- INEPs não encontrados são listados ao final"
        )

    st.markdown('<hr class="saas-divider">', unsafe_allow_html=True)
    st.markdown("#### Fontes adicionais (opcional)")

    col_om_bm, col_zy_bm = st.columns(2)
    with col_om_bm:
        st.caption("🔵 **Omada** — export do portal On Premise Systems")
        bm_omada = st.file_uploader("Export Omada (.xlsx)", type=["xlsx"], key="bm_omada")
        if bm_omada:
            df_bm_om = pd.read_excel(bm_omada)
            bm_omada.seek(0)
            st.caption(f"{len(df_bm_om)} controllers carregados")
    with col_zy_bm:
        st.caption("🟣 **Zyxel** — export CSV do Nebula (Overview → Sites)")
        bm_zyxel = st.file_uploader("Export Zyxel (.csv)", type=["csv"], key="bm_zyxel")
        if bm_zyxel:
            df_bm_zy = pd.read_csv(bm_zyxel)
            bm_zyxel.seek(0)
            st.caption(f"{len(df_bm_zy)} sites carregados")

    st.markdown('<hr class="saas-divider">', unsafe_allow_html=True)

    tem_fonte_bm = alvo or bm_omada or bm_zyxel
    if not tem_fonte_bm:
        st.info("Configure ao menos uma fonte: conta Ubiquiti na barra lateral, ou upload de export Omada/Zyxel.")

    if st.button("🔍 Buscar", type="primary", disabled=not tem_fonte_bm):
        if not txt.strip():
            st.warning("Insira ao menos um INEP.")
        else:
            ineps_m = {i.strip() for i in txt.replace("\n",",").split(",") if i.strip()}
            res_bm, erros = {}, []

            if alvo:
                with st.spinner(f"Consultando Ubiquiti ({len(alvo)} conta(s))..."):
                    for c in alvo:
                        try:
                            adicionar_candidatos_por_inep(
                                res_bm,
                                buscar_ineps_ubiquiti(c["api_key"], c["apelido"], ineps_m),
                            )
                        except Exception as e:
                            erros.append(f"Ubiquiti {c['apelido']}: {e}")

            if bm_omada:
                try:
                    bm_omada.seek(0)
                    res_om = processar_export_omada(pd.read_excel(bm_omada))
                    adicionar_candidatos_por_inep(res_bm, res_om, ineps_m)
                except Exception as e:
                    erros.append(f"Omada: {e}")

            if bm_zyxel:
                try:
                    bm_zyxel.seek(0)
                    res_zy = processar_export_zyxel(pd.read_csv(bm_zyxel))
                    adicionar_candidatos_por_inep(res_bm, res_zy, ineps_m)
                except Exception as e:
                    erros.append(f"Zyxel: {e}")

            for e in erros: st.error(e)

            # Deduplicação: mantém a melhor entrada por INEP
            res_bm = deduplicar_por_inep(res_bm)

            if res_bm:
                dfm = pd.DataFrame.from_dict(res_bm, orient="index").reset_index()
                dfm.rename(columns={"index": "INEP"}, inplace=True)
                col_c1, col_c2, col_c3, col_c4 = st.columns(4)
                col_c1.metric("Encontrados",     len(dfm))
                col_c2.metric("Online",  dfm["Status Rede"].str.contains("ONLINE",  na=False).sum())
                col_c3.metric("Offline", dfm["Status Rede"].str.contains("OFFLINE", na=False).sum())
                col_c4.metric("Não encontrados", len(ineps_m) - len(res_bm))
                st.dataframe(dfm.style.map(cor_status, subset=["Status Rede"]), use_container_width=True)
                nao = ineps_m - set(res_bm.keys())
                if nao:
                    st.warning(f"{len(nao)} não encontrado(s): `{', '.join(sorted(nao))}`")
                st.download_button("⬇️ Exportar resultado", to_xlsx(dfm),
                    f"busca_manual_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.error("Nenhum INEP localizado em nenhuma das fontes consultadas.")

# ─────────────────────────────────────────
# ABA 4 — RAIO-X
# ─────────────────────────────────────────
with t4:
    if not alvo:
        st.warning("Selecione pelo menos uma conta na barra lateral.")
    else:
        c1, c2 = st.columns([1, 2])
        with c1:
            conta_rx = st.selectbox("Conta:", [c["apelido"] for c in alvo])
            filtro   = st.text_input("Filtrar por nome/INEP:", placeholder="Ex: FEIJÓ ou 12131229")
            btn      = st.button("🔬 Sondar", type="primary", use_container_width=True)
        with c2:
            if btn:
                key_rx = next(c["api_key"] for c in alvo if c["apelido"] == conta_rx)
                with st.spinner(f"Carregando hosts de '{conta_rx}'..."):
                    try:
                        hosts = get_paginated_hosts(key_rx)
                        rows  = []
                        for h in hosts:
                            d = extrair_host(h)
                            rows.append({
                                "Nome": d["nome"],
                                "Estado": d["estado"],
                                LAST_SEEN_COLUMN: d["ultimo_sinal"],
                                "IP": d["ip"],
                                "ISP": d["isp"],
                                "Uptime": d["uptime"],
                            })
                        rx_columns = ["Nome", "Estado", LAST_SEEN_COLUMN, "IP", "ISP", "Uptime"]
                        df_rx = pd.DataFrame(rows, columns=rx_columns)
                        if filtro:
                            df_rx = df_rx[df_rx["Nome"].str.upper().str.contains(filtro.upper(), na=False)]
                        m1, m2, m3 = st.columns(3)
                        m1.metric("Total hosts", len(df_rx))
                        m2.metric("Conectados",  (df_rx["Estado"] == "connected").sum())
                        m3.metric("Offline",     (df_rx["Estado"] == "disconnected").sum())
                        if df_rx.empty:
                            st.info("Nenhum host encontrado para esta conta/filtro.")
                        else:
                            st.dataframe(df_rx.style.map(cor_estado, subset=["Estado"]),
                                         use_container_width=True, height=440)
                        st.download_button("⬇️ Exportar", to_xlsx(df_rx),
                            f"raio_x_{conta_rx}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    except Exception as e:
                        st.error(f"Erro: {e}")
            else:
                st.info("Selecione uma conta e clique em 'Sondar'.")

# ─────────────────────────────────────────
# ABA 5 — DIAGNÓSTICO
# ─────────────────────────────────────────
with t5:
    if not alvo:
        st.warning("Selecione pelo menos uma conta na barra lateral.")
    else:
        if st.button("🩺 Testar Todas as Contas Ubiquiti", type="primary"):
            for c in alvo:
                with st.spinner(f"Testando {c['apelido']}..."):
                    ok, msg = testar_conta(c["api_key"])
                if ok: st.success(f"**{c['apelido']}** — {msg}")
                else:  st.error(f"**{c['apelido']}** — {msg}")

        st.markdown("---")
        st.markdown("**Contas selecionadas:**")
        for c in alvo:
            k    = c["api_key"]
            prev = k[:6] + "•••" + k[-4:] if len(k) > 10 else "⚠️ inválida"
            st.markdown(f"- **{c['apelido']}** — `{prev}`")

# ─────────────────────────────────────────
# ABA 6 — AJUDA
# ─────────────────────────────────────────
with t6:
    st.markdown("### Como usar o Hub Redes — EACE")
    st.markdown("""
**Relatório Diário:**
1. Acesse **Área Restrita > Fechamento Diário** na barra lateral
2. Informe a senha da área restrita
3. Preencha os dados operacionais do dia e envie a planilha .xlsx
4. Clique em **Concluir Fechamento**
5. Acesse **Relatório Diário** para visualizar e baixar o relatório

**Inventário Geral:**
- Coleta TODAS as escolas de todas as plataformas, sem depender de chamados
- Ubiquiti: buscado automaticamente via API (todas as contas ativas)
- Omada e Zyxel: faça upload dos exports na aba Inventário
- Campo de pesquisa por INEP ou nome da escola
- Exportação filtrada ou completa em .xlsx

**Como exportar o Zyxel:**
Zyxel Nebula → Overview → aba Sites → ícone de download (CSV) no canto superior direito.

**Como exportar o Omada:**
`use1-omada-cloud.tplinkcloud.com` → On Premise Systems → botão **Export**.

**Legenda de cores:**
- 🟢 UBIQUITI - ONLINE / 🔴 UBIQUITI - OFFLINE
- 🔵 OMADA - ONLINE / 🟠 OMADA - OFFLINE
- 🟣 ZYXEL - ONLINE / 🩷 ZYXEL - OFFLINE / 🟡 ZYXEL - ALERTA
- ⬛ NÃO ENCONTRADO
    """)
